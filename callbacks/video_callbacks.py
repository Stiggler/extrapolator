from dash import Output, Input, State, dcc, callback_context, ctx, dash_table
import dash
import pandas as pd
import sqlite3
import re
from helpers import decimal_to_hms
from helpers import convert_timedelta_to_decimal



def register_video_callbacks(app):

    @app.callback(
        [
            Output("percentages-status", "children"),
            Output("percentages-table", "data"),
            Output("percentages-table", "columns"),
            Output("field-selector", "options")
        ],
        Input("calculate-percentages", "n_clicks"),
        State("mm-dimensions", "value"),
        State("ea-dimensions", "value")
    )
    def calculate_percentages(n_clicks, mm_dims, ea_dims):
        if not n_clicks:
            return "", [], [], []
        group_by_cols = (mm_dims or []) + (ea_dims or [])
        if not group_by_cols:
            return "Bitte wählen Sie mindestens eine Dimension aus.", [], []

        group_by_clause = ", ".join(group_by_cols)
        join_condition = " AND ".join([f"v2.{dim} = v.{dim}" for dim in mm_dims]) if mm_dims else "1=1"

        query = f"""
        SELECT 
             {group_by_clause},
             COUNT(DISTINCT bid) AS count_bid,
             SUM(mentions) AS sum_mentions,
             SUM(visibility) AS sum_visibility,
             (
               SELECT SUM(CASE WHEN tool IS NULL OR tool = '' THEN broadcasting_time ELSE 0 END)
               FROM video AS v2
               WHERE v2.hr_basis = 'Basis' AND {join_condition}
             ) AS sum_broadcasting_time
        FROM video AS v
        WHERE hr_basis = 'Basis'
        GROUP BY {group_by_clause}
        """
        if ea_dims:
            having_conditions = " AND ".join([f"({dim} IS NULL OR {dim} = '')" for dim in ea_dims])
            query += f"\nHAVING NOT ({having_conditions})"

        conn = sqlite3.connect("data.db")
        df_raw = pd.read_sql(query, conn)
        conn.close()
        if df_raw.empty:
            return "Keine Daten gefunden.", [], [], []


        df = df_raw.copy()
        df['sum_mentions'] = df['sum_mentions'].fillna(0).astype(int)
        df['sum_visibility_raw'] = df['sum_visibility']
        df['sum_broadcasting_time_raw'] = df['sum_broadcasting_time']
        df['sum_visibility'] = df['sum_visibility_raw'].apply(decimal_to_hms)
        df['sum_broadcasting_time'] = df['sum_broadcasting_time_raw'].apply(decimal_to_hms)
        df['visibility_share'] = df.apply(
            lambda row: f"{(row['sum_visibility_raw'] / row['sum_broadcasting_time_raw'] * 100):.2f}%"
            if row['sum_broadcasting_time_raw'] else "N/A", axis=1
        )
        df['avg_mention'] = df.apply(
            lambda row: row['sum_visibility_raw'] / row['sum_mentions'] if row['sum_mentions'] != 0 else 0, axis=1
        )
        df['avg_mention'] = df['avg_mention'].apply(decimal_to_hms)

        final_cols = group_by_cols + [
            "sum_mentions", "avg_mention", "sum_visibility", "sum_broadcasting_time", "visibility_share"
        ]
        final_df = df[final_cols]
        columns = [{"name": col, "id": col, "editable": col in ["visibility_share", "avg_mention"]} for col in final_df.columns]
        data = final_df.to_dict("records")

        conn = sqlite3.connect("data.db")
        final_df.to_sql("percent", conn, if_exists="replace", index=False)
        conn.close()

        field_options = [{"label": col["name"], "value": col["id"]} for col in columns]
        return "Berechnung erfolgreich.", data, columns, field_options




    @app.callback(
        Output("extrapolate-status", "children"),
        Input("extrapolate", "n_clicks"),
        State("mm-dimensions", "value"),
        State("ea-dimensions", "value"),
        prevent_initial_call=True
    )
    def extrapolate_hr(n_clicks, mm_dims, ea_dims):
        if not n_clicks:
            return dash.no_update

        # Dimensionen initialisieren
        mm_dims = mm_dims or []
        ea_dims = ea_dims or []
        group_by_cols = mm_dims + ea_dims

        # Daten aus Datenbank laden
        conn = sqlite3.connect("data.db")
        df_video = pd.read_sql("SELECT * FROM video WHERE hr_basis = 'HR'", conn)
        df_percent = pd.read_sql("SELECT * FROM percent", conn)
        conn.close()

        # Whitespace-Bereinigung für alle group_by_cols
        for col in group_by_cols:
            if col in df_video.columns:
                df_video[col] = df_video[col].astype(str).str.strip()
            if col in df_percent.columns:
                df_percent[col] = df_percent[col].astype(str).str.strip()

        # 1) Filter auf MM-Dimensionen: nur diejenigen Prozent-Zeilen behalten, die in df_video vorkommen
        if mm_dims:
            valid_mm = df_video[mm_dims].drop_duplicates()
            df_percent = df_percent.merge(valid_mm, on=mm_dims, how="inner")

        # 2) avg_mention konvertieren
        if "avg_mention" in df_percent.columns:
            df_percent["avg_mention"] = pd.to_timedelta(df_percent["avg_mention"], errors="coerce")
            df_percent["avg_mention_numeric"] = df_percent["avg_mention"].apply(convert_timedelta_to_decimal)

        # 3) Sichtbarkeit im Percent-DF neu berechnen
        if "visibility" in df_percent.columns:
            df_percent.drop(columns=["visibility"], inplace=True)
        if "visibility_share" in df_percent.columns and "sum_broadcasting_time" in df_percent.columns:
            # Visibility-Share in Float
            df_percent["visibility_share_float"] = (
                df_percent["visibility_share"].str.replace("%", "").str.replace(",", ".").astype(float) / 100
            )
            # sum_broadcasting_time (HH:MM:SS) in Tage umwandeln
            df_percent["sum_broadcasting_time_days"] = (
                pd.to_timedelta(df_percent["sum_broadcasting_time"], errors="coerce").dt.total_seconds() / 86400
            )
            # Absolute Visibility pro Kombination (in Tagen)
            df_percent["visibility"] = (
                df_percent["visibility_share_float"] * df_percent["sum_broadcasting_time_days"]
            )

        # 4) Merge auf MM-Dimensionen (EA-Zeilen werden dupliziert)) Merge auf MM-Dimensionen (EA-Zeilen werden dupliziert)) Merge auf MM-Dimensionen (EA-Zeilen werden dupliziert)
        try:
            df_merged = pd.merge(
                df_video,
                df_percent,
                on=mm_dims,
                how="inner",
                suffixes=("", "_percent")
            )
        except Exception as e:
            return f"❌ Merge-Fehler: {e}"

        if df_merged.empty:
            return "⚠️ Keine passenden Kombinationen zwischen HR-Zeilen und Prozentwerten gefunden."

        # 5) EA-Dimensionen aus df_percent übernehmen und Suffixe entfernen
        for col in ea_dims:
            percent_col = f"{col}_percent"
            if percent_col in df_merged.columns:
                df_merged[col] = df_merged[percent_col]
        drop_cols = [f"{col}_percent" for col in ea_dims]
        df_merged.drop(columns=drop_cols, inplace=True)

        # 6) Neue Sichtbarkeit berechnen: broadcasting_time * visibility_share_float
        try:
            df_merged["visibility"] = (
                df_merged["broadcasting_time"] * df_merged["visibility_share_float"]
            )
        except Exception:
            df_merged["broadcasting_time_num"] = (
                pd.to_timedelta(df_merged["broadcasting_time"], errors="coerce").dt.total_seconds() / 86400
            )
            df_merged["visibility"] = (
                df_merged["broadcasting_time_num"] * df_merged["visibility_share_float"]
            )

        # Neue Zeilen mit visibility == 0 entfernen
        df_merged = df_merged[df_merged["visibility"] > 0]

        # 7) Mentions berechnen basierend auf neuer visibility und avg_mention_numeric
        df_merged["mentions"] = df_merged.apply(
            lambda row: int(row["visibility"] / row["avg_mention_numeric"]) if pd.notnull(row["visibility"]) and row["avg_mention_numeric"] > 0 else 0,
            axis=1
        )
        # Falls visibility > 0 aber mentions == 0, setze mentions auf 1
        df_merged.loc[(df_merged["visibility"] > 0) & (df_merged["mentions"] == 0), "mentions"] = 1

    # 8) In Datenbank speichern
        conn = sqlite3.connect("data.db")
        df_merged.to_sql("video_final", conn, if_exists="replace", index=False)
        df_merged.to_sql("hr_bewegt", conn, if_exists="replace", index=False)
        conn.close()

        return f"✅ Extrapolation erfolgreich: {len(df_merged)} Zeilen gespeichert (hr_bewegt)."





    @app.callback(
        [Output("results-status", "children"),
         Output("results-table", "data"),
         Output("results-table", "columns")],
        [Input("calculate-results", "n_clicks"),
         Input("calculate-results2", "n_clicks")],
        [State("mm-dimensions-results-video", "value"),
         State("ea-dimensions-results-video", "value")]
    )
    def combined_results(n_clicks1, n_clicks2, mm_dims, ea_dims):
        ctx = dash.callback_context
        if not ctx.triggered:
            return "", [], []

        triggered_id = ctx.triggered[0]['prop_id'].split('.')[0]
        db_path = "data.db"

        if triggered_id == "calculate-results":
            conn = sqlite3.connect(db_path)
            df_video = pd.read_sql("SELECT * FROM video", conn)
            df_hr = pd.read_sql("SELECT * FROM hr_bewegt", conn)
            conn.close()

            if "sponsor_percent" in df_hr.columns:
                df_hr["sponsor"] = df_hr["sponsor_percent"]
                df_hr.drop(columns=["sponsor_percent"], inplace=True)
            if "tool_percent" in df_hr.columns:
                df_hr["tool"] = df_hr["tool_percent"]
                df_hr.drop(columns=["tool_percent"], inplace=True)

            df_final = pd.concat([df_video, df_hr], ignore_index=True)

            try:
                df_final["sponsoring_value_cpt"] = (df_final["visibility"] * df_final["reach"] * 86400 / 1000 * 10 * 1000000) / 30
                df_final["sponsorship_contacts"] = (df_final["visibility"] * df_final["reach"] * 86400 / 30)
                df_final["ave_100"] = (df_final["visibility"] * 86400) * (df_final["advertising_price_TV"] / 30)
                df_final["sponsoring_value_cpt"] = df_final["sponsoring_value_cpt"].fillna(0).apply(lambda x: int(x))
                df_final["ave_100"] = df_final["ave_100"].fillna(0).apply(lambda x: int(x))
            except Exception as e:
                return f"Fehler bei der Berechnung von sponsoring_value_cpt: {e}", [], []

            conn = sqlite3.connect(db_path)
            df_final.to_sql("video_final", conn, if_exists="replace", index=False)
            conn.close()

            return f"Neue Tabelle 'video_final' erstellt: {len(df_final)} Zeilen, Sponsoring_Value_CPT aktualisiert.", [], []

        elif triggered_id == "calculate-results2":
            group_by_cols = (mm_dims or []) + (ea_dims or [])
            if not group_by_cols:
                return "Bitte wählen Sie mindestens eine Dimension aus.", [], []

            conn = sqlite3.connect(db_path)
            df = pd.read_sql("SELECT * FROM video_final", conn)
            conn.close()

            if df.empty:
                return "Die Tabelle video_final ist leer.", [], []

            df = df[df['hr_basis'].isin(['Basis', 'HR'])]

            grouped_vis = df.groupby(group_by_cols + ['hr_basis'], as_index=False)['visibility'].sum()
            pivot_vis = grouped_vis.pivot_table(index=group_by_cols, columns='hr_basis', values='visibility', fill_value=0).reset_index()
            pivot_vis.rename(columns={'Basis': 'sum_visibility_basis', 'HR': 'sum_visibility_hr'}, inplace=True)

            grouped_extra = df.groupby(group_by_cols + ['hr_basis'], as_index=False).agg({
                'bid': pd.Series.nunique,
                'ave_100': 'sum'
            })
            pivot_bid = grouped_extra.pivot_table(index=group_by_cols, columns='hr_basis', values='bid', fill_value=0).reset_index()
            pivot_bid.rename(columns={'Basis': 'bid_count_basis', 'HR': 'bid_count_hr'}, inplace=True)
            pivot_ave = grouped_extra.pivot_table(index=group_by_cols, columns='hr_basis', values='ave_100', fill_value=0).reset_index()
            pivot_ave.rename(columns={'Basis': 'sum_ave_100_basis', 'HR': 'sum_ave_100_hr'}, inplace=True)

            final_df = pivot_vis.merge(pivot_bid, on=group_by_cols, how='outer') \
                                .merge(pivot_ave, on=group_by_cols, how='outer')

            final_df["sum_visibility_basis"] = final_df["sum_visibility_basis"].apply(decimal_to_hms)
            final_df["sum_visibility_hr"] = final_df["sum_visibility_hr"].apply(decimal_to_hms)
            final_df["bid_count_basis"] = final_df["bid_count_basis"].apply(lambda x: format(int(x), ",d"))
            final_df["bid_count_hr"] = final_df["bid_count_hr"].apply(lambda x: format(int(x), ",d"))
            final_df["sum_ave_100_basis"] = final_df["sum_ave_100_basis"].apply(lambda x: format(int(x), ",d"))
            final_df["sum_ave_100_hr"] = final_df["sum_ave_100_hr"].apply(lambda x: format(int(x), ",d"))

            columns = [{"name": col, "id": col} for col in final_df.columns]
            data = final_df.to_dict("records")
            return f"Ergebnisse berechnet: {len(final_df)} Gruppen gefunden.", data, columns

        return "", [], []


    from dash import dcc  # sicherstellen, dass dcc importiert ist
    from io import BytesIO
    import openpyxl
    from openpyxl.utils import get_column_letter

    @app.callback(
        Output("download", "data"),
        Input("export-button", "n_clicks"),
        prevent_initial_call=True
    )
    def export_video_to_excel(n_clicks):
        if not n_clicks:
            return None

        conn = sqlite3.connect("data.db")
        df = pd.read_sql("SELECT * FROM video_final", conn)
        conn.close()

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="data")
            worksheet = writer.sheets["data"]
            for i, col in enumerate(df.columns):
                col_letter = get_column_letter(i + 1)
                if col.lower() in ["pr_value", "ave_100", "ave_weighted", "sponsoring_value_cpt"]:
                    for cell in worksheet[col_letter][1:]:
                        cell.number_format = '#,##0'
                elif col.lower() in ["broadcasting_time", "visibility", "apt", "start_time_program", "end_time_program", "start_time_item"]:
                    for cell in worksheet[col_letter][1:]:
                        cell.number_format = 'h:mm:ss'
        output.seek(0)
        return dcc.send_bytes(output.getvalue(), "video_final.xlsx")



    @app.callback(
        Output("percentages-table", "data", allow_duplicate=True),
        Input("duplicate-percentages-rows", "n_clicks"),
        State("percentages-table", "data"),
        State("percentages-table", "columns"),
        State("percentages-table", "selected_rows"),
        prevent_initial_call=True
    )
    def modify_percentages_table(n_duplicate, data, columns, selected_rows):
        if not selected_rows or data is None:
            return data
        duplicated = [data[i].copy() for i in selected_rows]
        return data + duplicated


        if triggered == "add-percentages-row":
            new_row = {col["id"]: "" for col in columns}
            data.append(new_row)
        elif triggered == "duplicate-percentages-rows" and selected_rows:
            duplicated = [data[i].copy() for i in selected_rows]
            data += duplicated

        return data


    @app.callback(
        Output("percentages-table", "data", allow_duplicate=True),
        Input("apply-field-value", "n_clicks"),
        State("percentages-table", "data"),
        State("percentages-table", "selected_rows"),
        State("field-selector", "value"),
        State("field-value", "value"),
        prevent_initial_call=True
    )
    def apply_field_value(n_clicks, data, selected_rows, field, value):
        if not selected_rows or not field:
            return data

        for i in selected_rows:
            data[i][field] = value
        return data


    @app.callback(
        [
            Output("basecheck-status", "children"),
            Output("basecheck-table", "data"),
            Output("basecheck-table", "columns")
        ],
        Input("calculate-basecheck", "n_clicks"),
        State("mm-dimensions-basecheck", "value"),
        prevent_initial_call=True
    )
    def calculate_basecheck(n_clicks, dimensions):
        if not dimensions:
            return "Bitte wählen Sie mindestens eine Dimension aus.", [], []

        conn = sqlite3.connect("data.db")
        df = pd.read_sql("SELECT * FROM video", conn)
        conn.close()

        if df.empty or "hr_basis" not in df.columns:
            return "Keine gültigen Daten in 'video' gefunden.", [], []

        # Gruppierung nach Dimensionen + hr_basis
        grouped = df.groupby(dimensions + ["hr_basis"], as_index=False).agg({
            "bid": pd.Series.nunique,
            "visibility": "sum",
            "broadcasting_time": "sum"
        })

        # Pivotieren
        pivot_bid = grouped.pivot_table(index=dimensions, columns="hr_basis", values="bid", fill_value=0).add_prefix("distinct_bid_").reset_index()
        pivot_vis = grouped.pivot_table(index=dimensions, columns="hr_basis", values="visibility", fill_value=0).add_prefix("visibility_").reset_index()
        pivot_bt  = grouped.pivot_table(index=dimensions, columns="hr_basis", values="broadcasting_time", fill_value=0).add_prefix("broadcasting_time_").reset_index()

        # Zusammenführen
        df_final = pivot_bid.merge(pivot_vis, on=dimensions).merge(pivot_bt, on=dimensions)

        # Zeitfelder umwandeln
        for col in df_final.columns:
            if col.startswith("visibility_") or col.startswith("broadcasting_time_"):
                df_final[col] = df_final[col].apply(decimal_to_hms)

        columns = []
        for col in df_final.columns:
            if col.startswith("distinct_bid_"):
                basis = col.replace("distinct_bid_", "")
                columns.append({"name": ["distinct_bid", basis], "id": col})
            elif col.startswith("visibility_"):
                basis = col.replace("visibility_", "")
                columns.append({"name": ["visibility", basis], "id": col})
            elif col.startswith("broadcasting_time_"):
                basis = col.replace("broadcasting_time_", "")
                columns.append({"name": ["broadcasting_time", basis], "id": col})
            else:
                columns.append({"name": [col, ""], "id": col})

        data = df_final.to_dict("records")

        return f"{len(df_final)} Gruppen gefunden.", data, columns

    @app.callback(
        Output("percentages-table", "selected_rows", allow_duplicate=True),
        Input("select-all-visible-rows", "n_clicks"),
        State("percentages-table", "data"),
        prevent_initial_call=True
    )
    def select_all_rows(n_clicks, data):
        return list(range(len(data)))

    @app.callback(
        Output("percentages-table", "selected_rows", allow_duplicate=True),
        Input("deselect-all-rows", "n_clicks"),
        prevent_initial_call=True
    )
    def deselect_all_rows(n_clicks):
        return []

    def safe_decimal_to_hms(val):
        import pandas as pd
        if pd.isnull(val):
            return ""
        if isinstance(val, str) and ":" in val:
            return val
        if isinstance(val, pd.Timedelta):
            total_seconds = int(val.total_seconds())
        else:
            try:
                total_seconds = int(float(val) * 86400)
            except Exception:
                return ""
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:d}:{minutes:02d}:{seconds:02d}"

    @app.callback(
        Output("download-percentages", "data"),
        Input("export-percentages-button", "n_clicks"),
        State("percentages-table", "data"),
        prevent_initial_call=True
    )
    def export_percentages(n_clicks, data):
        if not data:
            return None
        import pandas as pd
        from io import BytesIO
        from openpyxl.utils import get_column_letter
        import openpyxl

        df = pd.DataFrame(data)

        # avg_mention als Zeitstring (Excel-Zeitformat)
        if "avg_mention" in df.columns:
            df["avg_mention"] = df["avg_mention"].apply(safe_decimal_to_hms)

        # visibility_share als Dezimalwert (Excel-Prozentwert)
        if "visibility_share" in df.columns:
            # Stelle sicher, dass alles float ist (und kein String wie „12,25 %“)
            # Wenn im DataFrame noch 12,25 steht, wandle um zu 0.1225
            df["visibility_share"] = (
                df["visibility_share"]
                .astype(str)
                .str.replace("%", "")
                .str.replace(",", ".")
                .astype(float)
                .apply(lambda x: x/100 if x > 1 else x)   # Nur falls versehentlich 12.25 statt 0.1225 im DataFrame steht
            )

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="percent")
            worksheet = writer.sheets["percent"]
            for i, col in enumerate(df.columns):
                col_letter = get_column_letter(i + 1)
                if col == "avg_mention":
                    for cell in worksheet[col_letter][1:]:
                        cell.number_format = "h:mm:ss"
                elif col == "visibility_share":
                    for cell in worksheet[col_letter][1:]:
                        cell.number_format = "0.00%"

        output.seek(0)
        return dcc.send_bytes(output.getvalue(), "percent.xlsx")






    def parse_excel_avg_mention(val):
        import pandas as pd
        if pd.isnull(val):
            return None
        # 1. Timedelta (Excel-Zeit als Intervall): in Tage wandeln
        if isinstance(val, pd.Timedelta):
            return val.total_seconds() / 86400
        # 2. Float/Int (Excel als Dezimal): direkt übernehmen
        if isinstance(val, (float, int)):
            return float(val)
        # 3. String: Ist es ein Dezimalwert (z. B. "0,0000925925" oder "0.0000925925")?
        val_str = str(val).replace(",", ".").strip()
        try:
            # Versuche, als Zahl zu interpretieren
            as_float = float(val_str)
            return as_float
        except Exception:
            pass
        # 4. Sonst: versuche, als Zeitstring zu parsen
        try:
            td = pd.to_timedelta(val_str)
            if pd.isnull(td):
                return None
            return td.total_seconds() / 86400
        except Exception:
            return None




    @app.callback(
        [
            Output("percentages-table", "data", allow_duplicate=True),
            Output("percentages-table", "columns", allow_duplicate=True),
            Output("field-selector", "options", allow_duplicate=True)
        ],
        Input("import-percentages-upload", "contents"),
        State("percentages-table", "data"),
        prevent_initial_call=True
    )
    def import_percentages(contents, existing_data):
        import base64
        import io
        import pandas as pd
        from helpers import decimal_to_hms

        def parse_excel_avg_mention(val):
            """Akzeptiert Excel-Zeit (float), Zeitstring oder Dezimal-Tage, gibt Dezimal-Tage zurück."""
            if pd.isnull(val):
                return None
            if isinstance(val, (float, int)):
                return float(val)
            # Excel-Zeit als Intervall/Timedelta
            if isinstance(val, pd.Timedelta):
                return val.total_seconds() / 86400
            try:
                td = pd.to_timedelta(val)
                if pd.isnull(td):
                    return None
                return td.total_seconds() / 86400
            except Exception:
                return None

        if contents is None:
            return dash.no_update

        # Datei dekodieren & einlesen
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)
        df_new = pd.read_excel(io.BytesIO(decoded), sheet_name="percent")

        # avg_mention in float (Dezimal-Tage) umwandeln (akzeptiert Timedelta, float, String)
        if "avg_mention" in df_new.columns:
            df_new["avg_mention"] = df_new["avg_mention"].apply(parse_excel_avg_mention)
            df_new["avg_mention"] = df_new["avg_mention"].apply(decimal_to_hms)


        # visibility_share als float (z.B. 0.2180)
        if "visibility_share" in df_new.columns:
            df_new["visibility_share"] = pd.to_numeric(df_new["visibility_share"], errors="coerce")
            # Immer als Prozent anzeigen (0.2516 wird zu "25.16%")
            df_new["visibility_share"] = df_new["visibility_share"].apply(
                lambda x: f"{x*100:.2f}%" if pd.notnull(x) else ""
            )




        # Alte Daten (falls vorhanden)
        df_existing = pd.DataFrame(existing_data) if existing_data else pd.DataFrame()

        # Gemeinsame Spalten bestimmen und zusammenführen
        if not df_existing.empty:
            common_cols = [col for col in df_existing.columns if col in df_new.columns]
            df_new = df_new[common_cols]
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        else:
            df_combined = df_new

        # Outputs
        data = df_combined.to_dict("records")
        columns = [{"name": col, "id": col, "editable": True} for col in df_combined.columns]
        dropdown_options = [{"label": col, "value": col} for col in df_combined.columns]

        return data, columns, dropdown_options





    @app.callback(
        Output("percentages-status", "children", allow_duplicate=True),
        Input("update-percentages", "n_clicks"),
        State("percentages-table", "data"),
        prevent_initial_call=True
    )
    def update_percentages_db(n_clicks, data):
        if not data:
            return "❌ Keine Daten zum Speichern."

        try:
            df = pd.DataFrame(data)
            conn = sqlite3.connect("data.db")
            df.to_sql("percent", conn, if_exists="replace", index=False)
            conn.close()
            return f"✅ Prozentwertetabelle erfolgreich gespeichert ({len(df)} Zeilen)."
        except Exception as e:
            return f"❌ Fehler beim Speichern: {e}"


    @app.callback(
        [
            Output("percentages-table", "data", allow_duplicate=True),
            Output("percentages-table", "selected_rows", allow_duplicate=True),
        ],
        Input("delete-percentages-rows", "n_clicks"),
        State("percentages-table", "data"),
        State("percentages-table", "selected_rows"),
        prevent_initial_call=True
    )
    def delete_percentages_rows(n_clicks, data, selected_rows):
        if not n_clicks or not data or not selected_rows:
            raise dash.exceptions.PreventUpdate

        data_new = [row for i, row in enumerate(data) if i not in selected_rows]

        return data_new, []
