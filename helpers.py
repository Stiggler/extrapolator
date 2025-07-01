import base64
import io
import os
import sqlite3
import pandas as pd
import datetime
from openpyxl import load_workbook

PARQUET_CACHE = "cache/latest_upload.parquet"

def read_time_column_with_openpyxl(path_or_buffer, column_name):
    wb = load_workbook(path_or_buffer, data_only=True)
    ws = wb["data"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    target_index = None
    for i, name in enumerate(header):
        if name and str(name).strip().lower() == column_name.strip().lower():
            target_index = i
            break
    values = []
    if target_index is not None:
        for row in ws.iter_rows(min_row=2):
            val = row[target_index].value
            if isinstance(val, datetime.timedelta):
                values.append(val)
            elif isinstance(val, (int, float)):
                values.append(pd.to_timedelta(val, unit="D"))
            elif isinstance(val, str):
                try:
                    values.append(pd.to_timedelta(val))
                except:
                    values.append(pd.NaT)
            elif hasattr(val, "hour") and hasattr(val, "minute") and hasattr(val, "second"):
                try:
                    values.append(pd.to_timedelta(f"{val.hour}:{val.minute}:{val.second}"))
                except:
                    values.append(pd.NaT)
            else:
                values.append(pd.NaT)
    return pd.Series(values)


def convert_timedelta_to_decimal(td):
    if pd.isnull(td):
        return None
    return td.total_seconds() / 86400


def decimal_to_hms(decimal_val):
    if pd.isnull(decimal_val):
        return ""
    total_seconds = decimal_val * 86400
    total_seconds = int(round(total_seconds))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

def time_to_timedelta(val):
    if pd.isna(val):
        return pd.NaT
    if isinstance(val, pd.Timedelta):
        return val
    if isinstance(val, (int, float)):
        # Excel-Zeitwert als Bruchteil eines Tages
        return pd.to_timedelta(val, unit="D")
    if isinstance(val, str):
        try:
            return pd.to_timedelta(val)
        except:
            return pd.NaT
    if hasattr(val, "hour") and hasattr(val, "minute") and hasattr(val, "second"):
        try:
            return pd.to_timedelta(f"{val.hour}:{val.minute}:{val.second}")
        except:
            return pd.NaT
    return pd.NaT





def parse_contents(contents, filename):
    """
    1) Nur ein einziger read_excel-Aufruf statt Zeile-für-Zeile mit Openpyxl.
    2) Anschließend Time-Spalten vectorisiert mit pandas.to_timedelta parsen.
    """
    # 1. Payload dekodieren
    content_type, content_string = contents.split(',')
    decoded = base64.b64decode(content_string)
    excel_io = io.BytesIO(decoded)

    # 2. Komplett-Import aller Spalten in C-geschriebenem Code
    ID_COLUMNS = ["bid", "id"]  # ggf. erweitern!
    df = pd.read_excel(
        excel_io,
        sheet_name="data",
        engine="openpyxl",
        na_values=["", "NA", None]
    )
    for col in ID_COLUMNS:
        if col in df.columns:
            # Alles als String + "X" hinten dran (oder ein anderes seltenes Zeichen)
            df[col] = df[col].astype(str) 




    def _parse_mixed_time(val):
        if isinstance(val, (int, float)) and not pd.isna(val):
            return float(val)
        try:
            if isinstance(val, str) and val.replace(",", ".").replace(".", "", 1).isdigit():
                return float(val.replace(",", "."))
        except:
            pass
        try:
            td = pd.to_timedelta(val, errors="coerce")
            if pd.isnull(td):
                return None
            return td.total_seconds() / 86400
        except:
            return None

    for col in ["visibility", "broadcasting_time"]:
        if col in df.columns:
            df[col] = df[col].apply(_parse_mixed_time)

    

    # 4. Weitere Timedeltas (wenn nötig) auf dieselbe Weise
    for col in ["apt", "program_duration", "start_time_program", "end_time_program", "start_time_item"]:
        if col in df.columns:
            df[col] = (
                pd.to_timedelta(df[col], errors="coerce")
                  .dt.total_seconds() / 86400
            )

    return df


def update_database(df: pd.DataFrame, mode: str, first_file: bool):
    """
    1) SQLite PRAGMAs für Bulk-Inserts.
    2) to_sql im Batch-Modus mit method='multi' und chunksize.
    3) Parquet-Cache für spätere Leseläufe.
    """
    db_path = "data.db"
    # ensure cache folder exists
    os.makedirs(os.path.dirname(PARQUET_CACHE), exist_ok=True)

    # 1) Bulk-Insert mit Pragmas
    with sqlite3.connect(db_path, timeout=30) as conn:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        # conn.execute("PRAGMA locking_mode=EXCLUSIVE;")

        if first_file:
            how = "replace" if mode == "replace" else "append"
        else:
            how = "append"
            # aufteilen auf gemeinsame Spalten, wie gehabt
            existing = pd.read_sql("SELECT * FROM data LIMIT 1", conn).columns
            df = df.reindex(columns=existing, fill_value=pd.NA)

        # SQLite limitiert auf 999 Variablen pro INSERT → chunk size = floor(999 / ncols)
        ncols   = len(df.columns)
        max_vars = 999
        chunk   = max(1, max_vars // ncols)
        df.to_sql(
            "data",
            conn,
            if_exists=how,
            index=False,
            method="multi",
            chunksize=chunk
        )
    # 2) Parquet-Cache aktualisieren (Connection sauber schließen)
    try:
        with sqlite3.connect(db_path, timeout=30) as cache_conn:
            df_cache = pd.read_sql("SELECT * FROM data", cache_conn)
        df_cache.to_parquet(PARQUET_CACHE, index=False, engine="pyarrow")
    except Exception:
        pass



def load_data(refresh_from_db: bool = False) -> pd.DataFrame:
    """
    Zentrale Daten-Ladefunktion:
    - per Parquet, wenn vorhanden
    - sonst per SQL
    """
    if not refresh_from_db and os.path.exists(PARQUET_CACHE):
        try:
            return pd.read_parquet(PARQUET_CACHE, engine="pyarrow")
        except Exception:
            pass

    df = pd.read_sql("SELECT * FROM data", sqlite3.connect("data.db", timeout=30))
    try:
        df.to_parquet(PARQUET_CACHE, index=False, engine="pyarrow")
    except Exception:
        pass
    return df



def get_aggregated_data():
    db_path = 'data.db'
    conn = sqlite3.connect(db_path)
    query = """
    SELECT 
        TRIM(hr_basis) AS hr_basis,
        COUNT(DISTINCT bid) AS distinct_bid,
        SUM(visibility) AS sum_visibility,
        SUM(CASE WHEN tool IS NULL OR tool = '' THEN broadcasting_time ELSE 0 END) AS sum_broadcasting_time
    FROM data
    WHERE (media = 'TV/OTT' OR (media = 'Social Media' AND post_type IN ('Video')))
    GROUP BY TRIM(hr_basis);
    """
    df = pd.read_sql(query, conn)
    conn.close()
    if not df.empty:
        df['sum_visibility'] = df['sum_visibility'].apply(decimal_to_hms)
        df['sum_broadcasting_time'] = df['sum_broadcasting_time'].apply(decimal_to_hms)
    return df


def get_aggregated_data_opposite():
    db_path = 'data.db'
    conn = sqlite3.connect(db_path)
    query = """
    SELECT 
        TRIM(hr_basis) AS hr_basis,
        COUNT(DISTINCT bid) AS distinct_bid,
        SUM(mentions) AS sum_mentions
    FROM data
    WHERE media IN ('Print', 'Online', 'Social Media')
      AND (post_type IS NULL OR post_type = '' OR post_type NOT IN ('Video'))
    GROUP BY TRIM(hr_basis);
    """
    df = pd.read_sql(query, conn)
    conn.close()
    return df
