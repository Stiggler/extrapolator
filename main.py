import base64
import io
import sqlite3
import pandas as pd
import dash
from dash import Dash, dcc, html, dash_table, Input, Output, State
import math
import random 
import numpy as np
from io import BytesIO
import plotly.express as px
from dash.dash_table.Format import Format, Group
import openpyxl
from openpyxl.utils import get_column_letter

app = Dash(__name__, title="IRIS-Extrapolator")

# ---------------- Hilfsfunktionen ----------------

def convert_timedelta_to_decimal(td):
    """Konvertiert einen Timedelta in den Anteil eines Tages.
       Falls td NaT ist, wird None zurückgegeben."""
    if pd.isnull(td):
        return None
    return td.total_seconds() / 86400

def decimal_to_hms(decimal_val):
    """
    Konvertiert einen Dezimalwert (Bruchteil eines Tages) in das Format h:mm:ss.
    Ist der Wert leer oder NaN, wird ein leerer String zurückgegeben.
    """
    if pd.isnull(decimal_val):
        return ""
    total_seconds = decimal_val * 86400
    total_seconds = int(round(total_seconds))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

# Umrechnung von h:mm:ss in einen Dezimalwert (Anteil eines Tages)
def hms_to_decimal(hms_str):
    try:
        parts = hms_str.split(':')
        if len(parts) != 3:
            return 0.0
        hours = float(parts[0])
        minutes = float(parts[1])
        seconds = float(parts[2])
        total_seconds = hours * 3600 + minutes * 60 + seconds
        return total_seconds / 86400
    except Exception:
        return 0.0

def parse_contents(contents, filename):
    """
    Dekodiert den Base64-String und liest mit pandas das Excel-Blatt "data" ein.
    Dabei werden broadcasting_time und visibility in Dezimalzahlen (Tagesanteil) umgewandelt,
    falls sie als Timedelta eingelesen wurden.
    """
    content_type, content_string = contents.split(',')
    decoded = base64.b64decode(content_string)
    try:
        df = pd.read_excel(
            io.BytesIO(decoded),
            sheet_name="data",
            engine="openpyxl",
            converters={
                'season': lambda x: str(int(x.toordinal() - pd.Timestamp("1899-12-30").toordinal()))
                        if isinstance(x, pd.Timestamp) else str(int(x)) if isinstance(x, (int, float)) else str(x)
            }
        )
        if 'broadcasting_time' in df.columns and pd.api.types.is_timedelta64_dtype(df['broadcasting_time']):
            df['broadcasting_time'] = df['broadcasting_time'].apply(convert_timedelta_to_decimal)
        if 'visibility' in df.columns and pd.api.types.is_timedelta64_dtype(df['visibility']):
            df['visibility'] = df['visibility'].apply(convert_timedelta_to_decimal)
        if 'apt' in df.columns and pd.api.types.is_timedelta64_dtype(df['apt']):
            df['apt'] = df['apt'].apply(convert_timedelta_to_decimal)            
        if 'program_duration' in df.columns and pd.api.types.is_timedelta64_dtype(df['program_duration']):
            df['program_duration'] = df['program_duration'].apply(convert_timedelta_to_decimal)           
        if 'start_time_program' in df.columns and pd.api.types.is_timedelta64_dtype(df['start_time_program']):
            df['start_time_program'] = df['start_time_program'].apply(convert_timedelta_to_decimal)  
        if 'end_time_program' in df.columns and pd.api.types.is_timedelta64_dtype(df['end_time_program']):
            df['end_time_program'] = df['end_time_program'].apply(convert_timedelta_to_decimal)  
        if 'start_time_item' in df.columns and pd.api.types.is_timedelta64_dtype(df['start_time_item']):
            df['start_time_item'] = df['start_time_item'].apply(convert_timedelta_to_decimal)
        if 'season' in df.columns:
            df['season'] = df['season'].apply(lambda x: str(int(x)) if isinstance(x, (float, int)) else str(x))
                           
        return df
    except Exception as e:
        print(f"Fehler beim Einlesen von {filename}: {e}")
        return None

def update_database(df, mode, first_file):
    """
    Schreibt den DataFrame in die SQLite-Datenbank "data.db".
    Bei der ersten Datei wird je nach Modus "replace" oder "append" verwendet.
    Für weitere Dateien wird immer angehängt.
    """
    db_path = 'data.db'
    conn = sqlite3.connect(db_path)
    if first_file:
        if_exists_option = "replace" if mode == "replace" else "append"
    else:
        if_exists_option = "append"
    df.to_sql("data", conn, if_exists=if_exists_option, index=False)
    conn.close()

def get_aggregated_data():
    """
    Aggregation für TV/OTT & Social Media (Video):
      - Filter: media = 'TV/OTT' oder (media = 'Social Media' und AND post_type IN ('Video', 'Story'))
      - Gruppierung nach hr_basis
      - Kennzahlen:
            • COUNT(DISTINCT bid) als distinct_bid
            • SUM(visibility) als sum_visibility
            • SUM(broadcasting_time) (nur für Zeilen, bei denen tool leer ist) als sum_broadcasting_time
    Anschließend werden die Zeitwerte ins Format h:mm:ss konvertiert.
    """
    db_path = 'data.db'
    conn = sqlite3.connect(db_path)
    query = """
    SELECT 
        TRIM(hr_basis) AS hr_basis,
        COUNT(DISTINCT bid) AS distinct_bid,
        SUM(visibility) AS sum_visibility,
        SUM(CASE WHEN tool IS NULL OR tool = '' THEN broadcasting_time ELSE 0 END) AS sum_broadcasting_time
    FROM data
    WHERE (media = 'TV/OTT' OR (media = 'Social Media' AND post_type IN ('Video', 'Story')))
    GROUP BY TRIM(hr_basis);
    """
    df = pd.read_sql(query, conn)
    conn.close()
    if not df.empty:
        df['sum_visibility'] = df['sum_visibility'].apply(decimal_to_hms)
        df['sum_broadcasting_time'] = df['sum_broadcasting_time'].apply(decimal_to_hms)
    return df

def get_aggregated_data_opposite():
    """
    Aggregation für Print, Online & Social Media (nicht Video):
      - Filter: media IN ('Print', 'Online', 'Social Media') und post_type NOT IN ('Video', 'Story')
      - Gruppierung nach hr_basis
      - Kennzahlen:
            • COUNT(DISTINCT bid) als distinct_bid
            • SUM(mentions) als sum_mentions
    """
    db_path = 'data.db'
    conn = sqlite3.connect(db_path)
    query = """
    SELECT 
        TRIM(hr_basis) AS hr_basis,
        COUNT(DISTINCT bid) AS distinct_bid,
        SUM(mentions) AS sum_mentions
    FROM data
    WHERE media IN ('Print', 'Online', 'Social Media')
      AND (post_type IS NULL OR post_type = '' OR post_type NOT IN ('Video', 'Story'))
        GROUP BY TRIM(hr_basis);
    """
    df = pd.read_sql(query, conn)
    conn.close()
    return df

# ---------------- Layout mit Tabs ----------------

app.layout = html.Div([
    dcc.Tabs([
        dcc.Tab(label="Import", selected_style={'backgroundColor': '#da8d00', 'color': 'white'}, children=[
            html.H1("Excel Daten Import Tool"),
            html.Div([
                html.Label("Import-Modus:"),
                dcc.RadioItems(
                    id="mode-radio",
                    options=[
                        {"label": "Append", "value": "append"},
                        {"label": "Replace", "value": "replace"}
                    ],
                    value="replace",
                    labelStyle={'display': 'inline-block', 'margin-right': '10px'}
                )
            ], style={'margin-bottom': '10px'}),
            dcc.Upload(
                id="upload-data",
                children=html.Div([
                    "Drag & Drop oder ",
                    html.A("Dateien auswählen")
                ]),
                multiple=True,
                accept=".xlsx",
                style={
                    'width': '100%',
                    'height': '60px',
                    'lineHeight': '60px',
                    'borderWidth': '1px',
                    'borderStyle': 'dashed',
                    'borderRadius': '5px',
                    'textAlign': 'center',
                    'color': 'white',
                    'background': '#73b8e6',
                    'margin-bottom': '20px'
                }
            ),
            html.Div(id="status", style={'margin-bottom': '20px'}),
            html.H2("Importübersicht"),
            html.Div([
                html.Div([
                    html.H3("Import Bewegtbild"),
                    dash_table.DataTable(
                        id="aggregated-table-1",
                        columns=[],
                        data=[],
                        style_table={'overflowX': 'auto'},
                        style_cell={'textAlign': 'left'},
                        style_header={
                            'backgroundColor': '#73b8e6',
                            'color': 'white'
                        }
                    )
                ], style={'width': '33%', 'display': 'inline-block', 'vertical-align': 'top', 'margin-right': '20px'}),
                html.Div([
                    html.H3("Import Nicht-Bewegtbild"),
                    dash_table.DataTable(
                        id="aggregated-table-2",
                        columns=[],
                        data=[],
                        style_table={'overflowX': 'auto'},
                        style_cell={'textAlign': 'left'},
                        style_header={
                            'backgroundColor': '#73b8e6',
                            'color': 'white'
                        }
                    )
                ], style={'width': '48%', 'display': 'inline-block', 'vertical-align': 'top'})
            ])
        ]),
        dcc.Tab(label="Bewegtbild", selected_style={'backgroundColor': '#da8d00', 'color': 'white'}, children=[
            dcc.Tabs(
                id="bewegtbild-subtabs",
                value="hochrechnung",  # Standardmäßig aktiver Untertab
                children=[
                    dcc.Tab(
                        label="Hochrechnung",
                        value="hochrechnung",
                        selected_style={'backgroundColor': '#73b8e6', 'color': 'white'},
                        children=[
                            html.H1("Bewegtbild – Hochrechnung"),
                            html.Br(),
                            html.Div([
                                html.Div([
                                    html.H3("MM-Dimensionen"),
                                    dcc.Dropdown(
                                        id='mm-dimensions',
                                        options=[
                                            {'label': 'media', 'value': 'media'},
                                            {'label': 'region', 'value': 'region'},
                                            {'label': 'country', 'value': 'country'},
                                            {'label': 'broadcaster', 'value': 'broadcaster'},
                                            {'label': 'channel', 'value': 'channel'},
                                            {'label': 'genre', 'value': 'genre'},
                                            {'label': 'sports', 'value': 'sports'},
                                            {'label': 'competition', 'value': 'competition'},
                                            {'label': 'season', 'value': 'season'},
                                            {'label': 'event', 'value': 'event'},
                                            {'label': 'venue', 'value': 'venue'},
                                            {'label': 'event_country', 'value': 'event_country'},
                                            {'label': 'post_type', 'value': 'post_type'},
                                            {'label': 'owned_channel', 'value': 'owned_channel'},
                                            {'label': 'discipline', 'value': 'discipline'},
                                            {'label': 'j1', 'value': 'j1'},
                                            {'label': 'j2', 'value': 'j2'},
                                            {'label': 'j3', 'value': 'j3'},
                                            {'label': 'j4', 'value': 'j4'},
                                            {'label': 'j5', 'value': 'j5'},
                                            {'label': 'hr1', 'value': 'hr1'},
                                            {'label': 'hr2', 'value': 'hr2'},
                                            {'label': 'hr3', 'value': 'hr3'},
                                            {'label': 'hr4', 'value': 'hr4'},
                                            {'label': 'hr5', 'value': 'hr5'}
                                        ],
                                        multi=True,
                                        placeholder="Wählen Sie MM-Dimensionen..."
                                    )
                                ], style={'width': '48%', 'display': 'inline-block', 'vertical-align': 'top', 'padding-right': '2%'}),
                                html.Div([
                                    html.H3("EA-Dimensionen"),
                                    dcc.Dropdown(
                                        id='ea-dimensions',
                                        options=[
                                            {'label': 'company', 'value': 'company'},
                                            {'label': 'sponsor', 'value': 'sponsor'},
                                            {'label': 'tool', 'value': 'tool'},
                                            {'label': 'personal_sponsorship', 'value': 'personal_sponsorship'},
                                            {'label': 'tool_location', 'value': 'tool_location'}
                                        ],
                                        multi=True,
                                        placeholder="Wählen Sie EA-Dimensionen..."
                                    )
                                ], style={'width': '48%', 'display': 'inline-block', 'vertical-align': 'top'})
                            ]),
                            html.Br(),
                            html.Button("Prozentwerte", id="calculate-percentages", style={'backgroundColor': 'green', 'color': 'white'}),
                            html.Div(id="percentages-status", style={'margin-top': '10px'}),
                            html.Br(),
                            html.Button("Extrapolate", id="extrapolate", style={'backgroundColor': 'purple', 'color': 'white'}),
                            html.Div(id="extrapolate-status", style={'margin-top': '10px'}),
                            html.Br(),
                            html.Button("Update Percentages", id="update-percentages", style={'backgroundColor': 'orange', 'color': 'white'}),
                            html.Div(id="update-percentages-status", style={'margin-top': '10px'}),
                            html.Br(),
                            dash_table.DataTable(
                                id="percentages-table",
                                columns=[],  # Dynamisch gesetzt
                                data=[],     # Dynamisch gesetzt
                                editable=True,
                                filter_action="native",
                                sort_action="native",
                                sort_mode="multi",
                                style_table={'overflowX': 'auto'},
                                style_cell={'textAlign': 'left'}
                            )
                        ]
                    ),
                    dcc.Tab(label="Ergebnisse",
                            selected_style={'backgroundColor': '#73b8e6', 'color': 'white'},
                            children=[
                        html.H1("Bewegtbild – Ergebnisse"),
                        html.Button("Berechne Ergebnisse", id="calculate-results", style={'backgroundColor': 'darkblue', 'color': 'white'}),
                        html.Button("Tabelle", id="calculate-results2", style={'backgroundColor': 'darkblue', 'color': 'white'}),
                        html.Button("Export", id="export-button", style={'backgroundColor': 'darkgreen', 'color': 'white', 'margin-left': '10px'}),
                        html.Div(id="results-status", style={'margin-top': '10px'}),
                        dcc.Download(id="download"),
                        dash_table.DataTable(
                            id="results-table",
                            columns=[],  # Wird dynamisch gesetzt
                            data=[],     # Wird dynamisch gesetzt
                            editable=False,
                            filter_action="native",
                            sort_action="native",
                            sort_mode="multi",
                            style_table={'overflowX': 'auto'},
                            style_cell={'textAlign': 'left'}
                        ),
                        html.Div("Platzhalter – hier folgen später Charts und weitere Visualisierungen.")
                    ])
,
                    dcc.Tab(
                        label="Basecheck",
                        value="basecheck",
                        selected_style={'backgroundColor': '#73b8e6', 'color': 'white'},
                        children=[
                            html.H1("Bewegtbild – Basecheck"),
                            html.Div([
                                html.H3("MM-Dimensionen (Basecheck)"),
                                dcc.Dropdown(
                                    id='mm-dimensions-basecheck',
                                    options=[
                                        {'label': 'media', 'value': 'media'},
                                        {'label': 'region', 'value': 'region'},
                                        {'label': 'country', 'value': 'country'},
                                        {'label': 'broadcaster', 'value': 'broadcaster'},
                                        {'label': 'channel', 'value': 'channel'},
                                        {'label': 'genre', 'value': 'genre'},
                                        {'label': 'sports', 'value': 'sports'},
                                        {'label': 'competition', 'value': 'competition'},
                                        {'label': 'season', 'value': 'season'},
                                        {'label': 'event', 'value': 'event'},
                                        {'label': 'venue', 'value': 'venue'},
                                        {'label': 'event_country', 'value': 'event_country'},
                                        {'label': 'post_type', 'value': 'post_type'},
                                        {'label': 'owned_channel', 'value': 'owned_channel'},
                                        {'label': 'discipline', 'value': 'discipline'},
                                        {'label': 'j1', 'value': 'j1'},
                                        {'label': 'j2', 'value': 'j2'},
                                        {'label': 'j3', 'value': 'j3'},
                                        {'label': 'j4', 'value': 'j4'},
                                        {'label': 'j5', 'value': 'j5'},
                                        {'label': 'hr1', 'value': 'hr1'},
                                        {'label': 'hr2', 'value': 'hr2'},
                                        {'label': 'hr3', 'value': 'hr3'},
                                        {'label': 'hr4', 'value': 'hr4'},
                                        {'label': 'hr5', 'value': 'hr5'}
                                    ],
                                    multi=True,
                                    placeholder="Wählen Sie MM-Dimensionen..."
                                )
                            ]),
                            html.Br(),
                            html.Button("Berechne Basecheck", id="calculate-basecheck", style={'backgroundColor': 'darkorange', 'color': 'white'}),
                            html.Div(id="basecheck-status", style={'margin-top': '10px'}),
                            html.Br(),
                            dash_table.DataTable(
                                id="basecheck-table",
                                columns=[],  # wird dynamisch gesetzt
                                data=[],     # wird dynamisch gesetzt
                                editable=False,
                                filter_action="native",
                                sort_action="native",
                                sort_mode="multi",
                                style_table={'overflowX': 'auto'},
                                style_cell={'textAlign': 'left'}
                            )
                        ]
                    )

                ]
            )
        ])
        ,
        dcc.Tab(label="Nicht-Bewegtbild", selected_style={'backgroundColor': '#da8d00', 'color': 'white'}, children=[
            dcc.Tabs(
                id="nicht-bewegtbild-subtabs",
                value="hochrechnung_nbv",  # Standardmäßig aktiver Untertab
                children=[
            dcc.Tab(
                label="Hochrechnung",
                value="hochrechnung_nbv",
                selected_style={'backgroundColor': '#73b8e6', 'color': 'white'},
                children=[
                    html.H1("Nicht-Bewegtbild – Hochrechnung"),
                    html.Div([
                        html.Div([
                            html.H3("MM-Dimensionen (Non-Video)"),
                            dcc.Dropdown(
                                id='mm-dimensions2',
                                options=[
                                    {'label': 'media', 'value': 'media'},
                                    {'label': 'region', 'value': 'region'},
                                    {'label': 'country', 'value': 'country'},
                                    {'label': 'broadcaster', 'value': 'broadcaster'},
                                    {'label': 'channel', 'value': 'channel'},
                                    {'label': 'genre', 'value': 'genre'},
                                    {'label': 'sports', 'value': 'sports'},
                                    {'label': 'competition', 'value': 'competition'},
                                    {'label': 'season', 'value': 'season'},
                                    {'label': 'event', 'value': 'event'},
                                    {'label': 'venue', 'value': 'venue'},
                                    {'label': 'event_country', 'value': 'event_country'},
                                    {'label': 'channel_type', 'value': 'channel_type'},
                                    {'label': 'post_type', 'value': 'post_type'},
                                    {'label': 'owned_channel', 'value': 'owned_channel'},
                                    {'label': 'discipline', 'value': 'discipline'},
                                    {'label': 'j1', 'value': 'j1'},
                                    {'label': 'j2', 'value': 'j2'},
                                    {'label': 'j3', 'value': 'j3'},
                                    {'label': 'j4', 'value': 'j4'},
                                    {'label': 'j5', 'value': 'j5'},
                                    {'label': 'hr1', 'value': 'hr1'},
                                    {'label': 'hr2', 'value': 'hr2'},
                                    {'label': 'hr3', 'value': 'hr3'},
                                    {'label': 'hr4', 'value': 'hr4'},
                                    {'label': 'hr5', 'value': 'hr5'}
                                ],
                                multi=True,
                                placeholder="Wählen Sie MM-Dimensionen..."
                            )
                        ], style={'width': '48%', 'display': 'inline-block', 'vertical-align': 'top', 'padding-right': '2%'}),
                        html.Div([
                            html.H3("EA-Dimensionen (Non-Video)"),
                            dcc.Dropdown(
                                id='ea-dimensions2',
                                options=[
                                    {'label': 'company', 'value': 'company'},
                                    {'label': 'sponsor', 'value': 'sponsor'},
                                    {'label': 'tool', 'value': 'tool'},
                                    {'label': 'personal_sponsorship', 'value': 'personal_sponsorship'},
                                    {'label': 'tool_location', 'value': 'tool_location'}
                                ],
                                multi=True,
                                placeholder="Wählen Sie EA-Dimensionen..."
                            )
                        ], style={'width': '48%', 'display': 'inline-block', 'vertical-align': 'top'})
                    ])

                    ,
                    html.Br(),
                    html.Button("Berechne Prozentwerte Non-Video", id="calculate-percentages2_nbv", style={'backgroundColor': 'green', 'color': 'white'}),
                    html.Div(id="nonvideo-percentages-status", style={'margin-top': '10px'}),
                    html.Button("Extrapolate Non-Video", id="extrapolate-nonvideo", style={'backgroundColor': 'purple', 'color': 'white'}),
                    html.Div(id="extrapolate-nonvideo-status", style={'margin-top': '10px'}),
                    html.Br(),
                    dash_table.DataTable(
                        id="nonvideo-percentages-table",
                        columns=[],  # wird dynamisch gesetzt
                        data=[],     # wird dynamisch gesetzt
                        editable=False,
                        filter_action="native",
                        sort_action="native",
                        sort_mode="multi",
                        style_table={'overflowX': 'auto'},
                        style_cell={'textAlign': 'left'}
                    )
                ]
            )

                    ,
                    dcc.Tab(
                        label="Ergebnisse",
                        value="ergebnisse_nbv",
                        selected_style={'backgroundColor': '#e74c3c', 'color': 'white'},
                        children=[
                            html.H1("Nicht-Bewegtbild – Ergebnisse"),
                            html.Div([
                                html.Div([
                                    html.H3("MM-Dimensionen Ergebnisse"),
                                    dcc.Dropdown(
                                        id='mm-dimensions-results',
                                        options=[
                                            {'label': 'media', 'value': 'media'},
                                            {'label': 'region', 'value': 'region'},
                                            {'label': 'country', 'value': 'country'},
                                            {'label': 'broadcaster', 'value': 'broadcaster'},
                                            {'label': 'channel', 'value': 'channel'},
                                            {'label': 'genre', 'value': 'genre'},
                                            {'label': 'sports', 'value': 'sports'},
                                            {'label': 'competition', 'value': 'competition'},
                                            {'label': 'season', 'value': 'season'},
                                            {'label': 'event', 'value': 'event'},
                                            {'label': 'venue', 'value': 'venue'},
                                            {'label': 'event_country', 'value': 'event_country'},
                                            {'label': 'j1', 'value': 'j1'},
                                            {'label': 'j2', 'value': 'j2'},
                                            {'label': 'j3', 'value': 'j3'},
                                            {'label': 'j4', 'value': 'j4'},
                                            {'label': 'j5', 'value': 'j5'},
                                            {'label': 'hr1', 'value': 'hr1'},
                                            {'label': 'hr2', 'value': 'hr2'},
                                            {'label': 'hr3', 'value': 'hr3'},
                                            {'label': 'hr4', 'value': 'hr4'},
                                            {'label': 'hr5', 'value': 'hr5'}
                                        ],
                                        multi=True,
                                        placeholder="Wählen Sie MM-Dimensionen für Ergebnisse..."
                                    )
                                ], style={'width': '48%', 'display': 'inline-block'}),
                                html.Div([
                                    html.H3("EA-Dimensionen Ergebnisse"),
                                    dcc.Dropdown(
                                        id='ea-dimensions-results',
                                        options=[
                                            {'label': 'company', 'value': 'company'},
                                            {'label': 'sponsor', 'value': 'sponsor'},
                                            {'label': 'tool', 'value': 'tool'},
                                            {'label': 'personal_sponsorship', 'value': 'personal_sponsorship'},
                                            {'label': 'tool_location', 'value': 'tool_location'},
                                        ],
                                        multi=True,
                                        placeholder="Wählen Sie EA-Dimensionen für Ergebnisse..."
                                    )
                                ], style={'width': '48%', 'display': 'inline-block'})
                            ], style={'margin-bottom': '20px'}),
                            # Neuer hr_basis-Filter
                            html.Div([
                                html.H3("Filter hr_basis"),
                                dcc.Dropdown(
                                    id='hr-basis-filter',
                                    options=[
                                        {'label': 'Alle', 'value': 'all'},
                                        {'label': 'Basis', 'value': 'Basis'},
                                        {'label': 'HR', 'value': 'HR'}
                                    ],
                                    value='all',
                                    clearable=False
                                )
                            ], style={'width': '30%', 'margin-bottom': '20px'}),
                            html.Button("Berechne Ergebnisse Non-Video", id="calculate-results-nonvideo", style={'backgroundColor': 'darkblue', 'color': 'white'}),
                            html.Div(id="nonvideo-results-status", style={'margin-top': '10px'}),
                            dash_table.DataTable(
                                id="nonvideo-results-table",
                                columns=[],  # Wird im Callback dynamisch gesetzt
                                data=[],     # Wird im Callback dynamisch gesetzt
                                editable=False,
                                filter_action="native",
                                sort_action="native",
                                sort_mode="multi",
                                style_table={'overflowX': 'auto'},
                                style_cell={'textAlign': 'left'}
                            ),
                            # Kreisdiagramm für die Verteilung von ave_weighted nach hr_basis
                            dcc.Graph(id="nonvideo-pie"),
                            html.Br(),
                            html.Button("Export Nicht-Bewegtbild", id="export-nonvideo-button", style={'backgroundColor': 'darkgreen', 'color': 'white', 'margin-left': '10px'}),
                            dcc.Download(id="download1"),
                            html.Div("Platzhalter – hier folgen weitere Visualisierungen.")
                        ]
                    )
,
                    dcc.Tab(
                        label="Basecheck",
                        value="basecheck_nbv",
                        selected_style={'backgroundColor': '#e74c3c', 'color': 'white'},
                        children=[
                            html.H1("Nicht-Bewegtbild – Basecheck"),
                            html.Div("Platzhalter – hier folgen Basecheck-Informationen für Nicht-Bewegtbild.")
                        ]
                    )
                ]
            )
        ])


            ])
        ])

# ---------------- Callback: Datenimport und Aggregation ----------------

@app.callback(
    [Output("status", "children"),
     Output("aggregated-table-1", "data"),
     Output("aggregated-table-1", "columns"),
     Output("aggregated-table-2", "data"),
     Output("aggregated-table-2", "columns")],
    Input("upload-data", "contents"),
    State("upload-data", "filename"),
    State("mode-radio", "value")
)
def update_on_upload(list_of_contents, list_of_names, mode):
    if list_of_contents is not None:
        status_messages = []
        first_file = True
        for contents, filename in zip(list_of_contents, list_of_names):
            if not filename.lower().startswith("report"):
                status_messages.append(f"Datei {filename} übersprungen (Name beginnt nicht mit 'report').")
                continue
            df = parse_contents(contents, filename)
            if df is None:
                status_messages.append(f"Fehler beim Verarbeiten von {filename}.")
                continue
            update_database(df, mode, first_file)
            first_file = False
            status_messages.append(f"Datei {filename} importiert.")
        
        # Automatische Erstellung der Tabelle "video"
        db_path = "data.db"
        conn = sqlite3.connect(db_path)
        query_video = """
            SELECT *
            FROM data
            WHERE (media = 'TV/OTT' OR (media = 'Social Media' AND post_type IN ('Video', 'Story')))
        """
        df_video = pd.read_sql(query_video, conn)
        conn.close()
        conn = sqlite3.connect(db_path)
        df_video.to_sql("video", conn, if_exists="replace", index=False)
        conn.close()
        status_messages.append(f"Tabelle 'video' in data.db erstellt: {len(df_video)} Zeilen wurden gespeichert.")
        
        # Automatische Erstellung der Tabelle "non_video"
        conn = sqlite3.connect(db_path)
        query_non_video = """
            SELECT *
            FROM data
            WHERE media IN ('Print', 'Online', 'Social Media')
              AND (post_type IS NULL OR post_type = '' OR post_type NOT IN ('Video', 'Story'))
        """
        df_non_video = pd.read_sql(query_non_video, conn)
        conn.close()
        conn = sqlite3.connect(db_path)
        df_non_video.to_sql("non_video", conn, if_exists="replace", index=False)
        conn.close()
        status_messages.append(f"Tabelle 'non_video' in data.db erstellt: {len(df_non_video)} Zeilen wurden gespeichert.")
        
        # Anschließend werden die aggregierten Daten geladen (wie bisher)
        df_agg1 = get_aggregated_data()
        if not df_agg1.empty:
            columns1 = [{"name": col, "id": col} for col in df_agg1.columns]
            data1 = df_agg1.to_dict("records")
        else:
            columns1 = []
            data1 = []
        df_agg2 = get_aggregated_data_opposite()
        if not df_agg2.empty:
            columns2 = [{"name": col, "id": col} for col in df_agg2.columns]
            data2 = df_agg2.to_dict("records")
        else:
            columns2 = []
            data2 = []
        
        return html.Div([html.Div(msg) for msg in status_messages]), data1, columns1, data2, columns2
    return "", [], [], [], []



# ---------------- Callback: Neue Tabelle "video" in data.db erstellen ----------------



# ---------------- Callback: Prozentwerte berechnen und in Tabelle "percent" speichern ----------------


@app.callback(
    [Output("percentages-status", "children"),
     Output("percentages-table", "data"),
     Output("percentages-table", "columns")],
    Input("calculate-percentages", "n_clicks"),
    State("mm-dimensions", "value"),
    State("ea-dimensions", "value")
)
def calculate_percentages(n_clicks, mm_dims, ea_dims):
    if not n_clicks:
        return "", [], []
    group_by_cols = []
    if mm_dims:
        group_by_cols.extend(mm_dims)
    if ea_dims:
        group_by_cols.extend(ea_dims)
    if not group_by_cols:
        return "Bitte wählen Sie mindestens eine Dimension aus.", [], []
    group_by_clause = ", ".join(group_by_cols)
    if mm_dims:
        join_condition = " AND ".join([f"v2.{dim} = v.{dim}" for dim in mm_dims])
    else:
        join_condition = "1=1"
    
    # Erweiterte Query: Zusätzlich wird COUNT(DISTINCT bid) als count_bid berechnet.
    query = f"""
    SELECT 
         {group_by_clause},
         COUNT(DISTINCT bid) AS count_bid,
         SUM(mentions) AS sum_mentions,
         SUM(visibility) AS sum_visibility,
         (
           SELECT SUM(CASE WHEN tool IS NULL OR tool = '' THEN broadcasting_time ELSE 0 END)
           FROM video AS v2
           WHERE v2.hr_basis = 'Basis'
             AND {join_condition}
         ) AS sum_broadcasting_time
    FROM video AS v
    WHERE hr_basis = 'Basis'
    GROUP BY {group_by_clause}
    """
    if ea_dims:
        having_conditions = " AND ".join([f"({dim} IS NULL OR {dim} = '')" for dim in ea_dims])
        query += f"\nHAVING NOT ({having_conditions})"
    
    db_path = "data.db"
    conn = sqlite3.connect(db_path)
    df_raw = pd.read_sql(query, conn)
    conn.close()
    
    if df_raw.empty:
        return "Keine Daten für die Berechnung gefunden.", [], []
    
    df = df_raw.copy()
    # Formatierung der Kennzahlen:
    df['sum_mentions'] = df['sum_mentions'].fillna(0).astype(int)
    df['sum_visibility_raw'] = df['sum_visibility']
    df['sum_broadcasting_time_raw'] = df['sum_broadcasting_time']
    df['sum_visibility'] = df['sum_visibility_raw'].apply(decimal_to_hms)
    df['sum_broadcasting_time'] = df['sum_broadcasting_time_raw'].apply(decimal_to_hms)
    df['visibility_share'] = df.apply(lambda row: f"{(row['sum_visibility_raw'] / row['sum_broadcasting_time_raw'] * 100):.2f}%"
                                      if row['sum_broadcasting_time_raw'] and row['sum_broadcasting_time_raw'] != 0 
                                      else "N/A", axis=1)
    # Berechnung von avg_mention = sum_visibility_raw / sum_mentions (ohne Rundung)
    df['avg_mention'] = df.apply(lambda row: row['sum_visibility_raw'] / row['sum_mentions'] 
                                  if row['sum_mentions'] != 0 else 0, axis=1)

    # Formatierung avg_mention als h:mm:ss (keine Rundung, alle Dezimalstellen beibehalten)
    df['avg_mention'] = df['avg_mention'].apply(decimal_to_hms)

    # Entferne count_bid aus dem finalen Output
    final_cols = group_by_cols + ["sum_mentions", "avg_mention", "sum_visibility", "sum_broadcasting_time", "visibility_share"]
    final_df = df[final_cols]
    columns = [{"name": col, "id": col, "editable": True if col in ["visibility_share", "avg_mention"] else False} for col in final_df.columns]
    data = final_df.to_dict("records")
    
    # Schreibe die Tabelle "percent" in die Datenbank:
    conn = sqlite3.connect(db_path)
    final_df.to_sql("percent", conn, if_exists="replace", index=False)
    conn.close()
    
    return "Berechnung erfolgreich.", data, columns


# ---------------- Callback: Aktualisierung der bearbeiteten Prozentwerte in der DB ----------------

@app.callback(
    Output("update-percentages-status", "children"),
    Input("update-percentages", "n_clicks"),
    State("percentages-table", "data")
)
def update_percentages_db(n_clicks, table_data):
    if not n_clicks:
        return ""
    df = pd.DataFrame(table_data)
    db_path = "data.db"
    conn = sqlite3.connect(db_path)
    df.to_sql("percent", conn, if_exists="replace", index=False)
    conn.close()
    return "Tabelle 'percent' wurde aktualisiert."

# ---------------- Callback: Hochrehnen ----------------

@app.callback(
    Output("extrapolate-status", "children"),
    Input("extrapolate", "n_clicks"),
    State("mm-dimensions", "value")
)
def extrapolate_hr(n_clicks, mm_dims):
    if not n_clicks:
        return ""
    if not mm_dims:
        return "Bitte wählen Sie mindestens eine MM-Dimension für die Hochrechnung aus."
    
    db_path = "data.db"
    conn = sqlite3.connect(db_path)
    # Lese alle HR-Zeilen aus der Tabelle video (hr_basis = 'HR')
    df_video = pd.read_sql("SELECT * FROM video WHERE hr_basis = 'HR'", conn)
    # Lese die Tabelle percent (enthält aggregierte Daten, inkl. visibility_share und avg_mention)
    df_percent = pd.read_sql("SELECT * FROM percent", conn)
    conn.close()
    
    if df_video.empty:
        return "Keine HR-Daten in der Tabelle video gefunden."
    if df_percent.empty:
        return "Keine Daten in der Tabelle percent gefunden."
    
    # Merge beider Tabellen anhand der ausgewählten MM-Dimensionen.
    # Dabei erhalten die Spalten aus video den Suffix "_video" und aus percent den Suffix "_percent"
    df_merged = pd.merge(df_video, df_percent, on=mm_dims, suffixes=("_video", "_percent"))
    
    # Überschreibe die Spalte visibility_share: Der Wert aus percent soll übernommen werden.
    df_merged["visibility_share"] = df_merged["visibility_share_percent"]
    
    # Hilfsfunktion zur Umrechnung eines Prozent-Strings in einen Faktor
    def convert_visibility_share(val):
        try:
            if isinstance(val, str) and "%" in val:
                return float(val.strip('%')) / 100.0
            elif isinstance(val, (int, float)):
                return float(val)
            else:
                return 0.0
        except Exception:
            return 0.0

    # Umrechnung: visibility_share als Faktor
    df_merged["visibility_share_factor"] = df_merged["visibility_share"].apply(convert_visibility_share)
    
    # Neue Berechnung: visibility = (visibility_share_factor) * (broadcasting_time aus video)
    # Hier verwenden wir den Wert aus der Spalte "broadcasting_time" (aus video)
    if "broadcasting_time" not in df_merged.columns:
        return "broadcasting_time aus video nicht gefunden."
    df_merged["visibility"] = df_merged["visibility_share_factor"] * df_merged["broadcasting_time"]
    
    # Wichtig: Da avg_mention in der percent-Tabelle als h:mm:ss formatiert vorliegt,
    # müssen wir diesen String wieder in einen numerischen Wert (Anteil eines Tages) umwandeln.
    def hms_to_decimal(hms_str):
        try:
            parts = hms_str.split(':')
            if len(parts) != 3:
                return 0.0
            hours = float(parts[0])
            minutes = float(parts[1])
            seconds = float(parts[2])
            total_seconds = hours * 3600 + minutes * 60 + seconds
            return total_seconds / 86400
        except Exception:
            return 0.0

    # Konvertiere die avg_mention-Spalte in einen numerischen Wert
    df_merged["avg_mention_numeric"] = df_merged["avg_mention"].apply(hms_to_decimal)
    
    # Berechne neue Mentions: mentions = visibility / avg_mention_numeric
    df_merged["mentions"] = df_merged["visibility"] / df_merged["avg_mention_numeric"]
    # Konvertiere in eine Ganzzahl; falls Ergebnis kleiner als 1, setze auf 1
    df_merged["mentions"] = df_merged["mentions"].apply(lambda x: int(x) if x >= 1 else 1)
    
    # Optional: Entferne Hilfsspalten, z. B. visibility_share_factor, avg_mention_numeric und visibility_share_percent
    df_merged.drop(columns=["visibility_share_factor", "avg_mention_numeric", "visibility_share_percent"], inplace=True)
    
    # Schreibe das Ergebnis in die Tabelle "hr_bewegt" in der Datenbank
    conn = sqlite3.connect(db_path)
    df_merged.to_sql("hr_bewegt", conn, if_exists="replace", index=False)
    conn.close()
    
    return f"Extrapolation abgeschlossen: {len(df_merged)} Zeilen wurden in 'hr_bewegt' erstellt."

# ---------------- Ergebnisse ----------------


@app.callback(
    [Output("results-status", "children"),
     Output("results-table", "data"),
     Output("results-table", "columns")],
    [Input("calculate-results", "n_clicks"),
     Input("calculate-results2", "n_clicks")],
    [State("mm-dimensions", "value"),
     State("ea-dimensions", "value")]
)
def combined_results(n_clicks1, n_clicks2, mm_dims, ea_dims):
    ctx = dash.callback_context
    if not ctx.triggered:
        return "", [], []
    
    triggered_id = ctx.triggered[0]['prop_id'].split('.')[0]
    db_path = "data.db"
    
    if triggered_id == "calculate-results":
        # Logik zum Erstellen der Tabelle "video_final" (Union von video und hr_bewegt)
        conn = sqlite3.connect(db_path)
        df_video = pd.read_sql("SELECT * FROM video", conn)
        df_hr = pd.read_sql("SELECT * FROM hr_bewegt", conn)
        conn.close()
        
        # In hr_bewegt sollen sponsor_percent und tool_percent als sponsor bzw. tool übernommen werden.
        if "sponsor_percent" in df_hr.columns:
            df_hr["sponsor"] = df_hr["sponsor_percent"]
            df_hr.drop(columns=["sponsor_percent"], inplace=True)
        if "tool_percent" in df_hr.columns:
            df_hr["tool"] = df_hr["tool_percent"]
            df_hr.drop(columns=["tool_percent"], inplace=True)
        
        # Vereine beide DataFrames (Union-All)
        df_final = pd.concat([df_video, df_hr], ignore_index=True)
        
        # Berechne sponsoring_value_cpt:
        # Formel: ((visibility * reach * 86400)/1000 * 10 * 1000000)/30
        try:
            df_final["sponsoring_value_cpt"] = (df_final["visibility"] * df_final["reach"] * 86400 / 1000 * 10 * 1000000) / 30
            df_final["sponsorship_contacts"] = (df_final["visibility"] * df_final["reach"] * 86400 / 30) 
            df_final["ave_100"] = (df_final["visibility"]* 86400) * (df_final["advertising_price_TV"] / 30)
            # Umwandlung in Integer (falls NaN, setze 0)
            df_final["sponsoring_value_cpt"] = df_final["sponsoring_value_cpt"].fillna(0).apply(lambda x: int(x))
            df_final["ave_100"] = df_final["ave_100"].fillna(0).apply(lambda x: int(x))
        except Exception as e:
            return f"Fehler bei der Berechnung von sponsoring_value_cpt: {e}", [], []
        
        # Schreibe den konsolidierten DataFrame in die Tabelle "video_final"
        conn = sqlite3.connect(db_path)
        df_final.to_sql("video_final", conn, if_exists="replace", index=False)
        conn.close()
        
        return f"Neue Tabelle 'video_final' erstellt: {len(df_final)} Zeilen, Sponsoring_Value_CPT aktualisiert.", [], []
    
    elif triggered_id == "calculate-results2":
        # Erstelle die Gruppierung: Nutze alle in mm- und ea-Dimensionen ausgewählten Felder
        group_by_cols = []
        if mm_dims:
            group_by_cols.extend(mm_dims)
        if ea_dims:
            group_by_cols.extend(ea_dims)
        if not group_by_cols:
            return "Bitte wählen Sie mindestens eine Dimension aus.", [], []
        
        conn = sqlite3.connect(db_path)
        df = pd.read_sql("SELECT * FROM video_final", conn)
        conn.close()
        
        if df.empty:
            return "Die Tabelle video_final ist leer.", [], []
        
        # Filtere nur Zeilen mit hr_basis "Basis" oder "HR"
        df = df[df['hr_basis'].isin(['Basis', 'HR'])]
        
        # Gruppierung für visibility: Berechne die Summe der visibility pro Gruppe und hr_basis
        grouped_vis = df.groupby(group_by_cols + ['hr_basis'], as_index=False)['visibility'].sum()
        pivot_vis = grouped_vis.pivot_table(index=group_by_cols, 
                                            columns='hr_basis', 
                                            values='visibility', 
                                            fill_value=0).reset_index()
        pivot_vis.rename(columns={'Basis': 'sum_visibility_basis', 'HR': 'sum_visibility_hr'}, inplace=True)
        
        # Gruppierung für die zusätzlichen Kennzahlen: 
        # - bid_count: Anzahl eindeutiger bid 
        # - sum_ave_100: Summe der Spalte ave_100
        grouped_extra = df.groupby(group_by_cols + ['hr_basis'], as_index=False).agg({
            'bid': pd.Series.nunique,
            'ave_100': 'sum'
        })
        # Pivotiere bid_count
        pivot_bid = grouped_extra.pivot_table(index=group_by_cols, 
                                            columns='hr_basis', 
                                            values='bid', 
                                            fill_value=0).reset_index()
        pivot_bid.rename(columns={'Basis': 'bid_count_basis', 'HR': 'bid_count_hr'}, inplace=True)
        # Pivotiere sum_ave_100
        pivot_ave = grouped_extra.pivot_table(index=group_by_cols, 
                                            columns='hr_basis', 
                                            values='ave_100', 
                                            fill_value=0).reset_index()
        pivot_ave.rename(columns={'Basis': 'sum_ave_100_basis', 'HR': 'sum_ave_100_hr'}, inplace=True)

        # Merge der Ergebnisse:
        final_df = pivot_vis.merge(pivot_bid, on=group_by_cols, how='outer') \
                            .merge(pivot_ave, on=group_by_cols, how='outer')

        final_df["sum_visibility_basis"] = final_df["sum_visibility_basis"].apply(decimal_to_hms)
        final_df["sum_visibility_hr"] = final_df["sum_visibility_hr"].apply(decimal_to_hms)

        # Formatierung: Umwandlung in Ganzzahlen mit Tausendertrennzeichen
        final_df["bid_count_basis"] = final_df["bid_count_basis"].apply(lambda x: format(int(x), ",d"))
        final_df["bid_count_hr"] = final_df["bid_count_hr"].apply(lambda x: format(int(x), ",d"))
        final_df["sum_ave_100_basis"] = final_df["sum_ave_100_basis"].apply(lambda x: format(int(x), ",d"))
        final_df["sum_ave_100_hr"] = final_df["sum_ave_100_hr"].apply(lambda x: format(int(x), ",d"))

        columns = [{"name": col, "id": col} for col in final_df.columns]
        data = final_df.to_dict("records")
        
        return f"Ergebnisse berechnet: {len(final_df)} Gruppen gefunden.", data, columns



    return "", [], []




# ---------------- Excel export video  ----------------

@app.callback(
    Output("download", "data"),
    Input("export-button", "n_clicks"),
    prevent_initial_call=True
)
def export_to_excel(n_clicks):
    if not n_clicks:
        return None
    db_path = "data.db"
    conn = sqlite3.connect(db_path)
    df = pd.read_sql("SELECT * FROM video_final", conn)
    conn.close()
    
    output = BytesIO()
    # Verwende den ExcelWriter als Kontextmanager
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="VideoFinal")
        workbook = writer.book
        worksheet = writer.sheets["VideoFinal"]
        
        # Durchlaufe alle Spalten und wende die gewünschte Formatierung an:
        for i, col in enumerate(df.columns):
            col_letter = get_column_letter(i+1)
            # Für "reach": Zahl mit 2 Dezimalstellen
            if col.lower() == ["reach", "sponsorship_contacts", "ratings_14+", "tv_ratings_14+"]:
                for cell in worksheet[col_letter][1:]:  # Überspringe den Header
                    cell.number_format = '0.00'
            # Für "broadcasting_time", "visibility" und "apt": Format h:mm:ss
            elif col.lower() in ["broadcasting_time", "visibility", "apt", "start_time_program", "end_time_program", "start_time_item"]:
                for cell in worksheet[col_letter][1:]:
                    cell.number_format = 'h:mm:ss'
             # Für sponsoring_value_cpt und pr_value: Zahl mit Tausendertrennzeichen (Punkt als Trenner) ohne Dezimalstellen
            elif col.lower() in ["sponsoring_value_cpt", "pr_value", "advertising_price_TV", "advertising_price_OTT", "ave_100", "ave_weighted"]:
                for cell in worksheet[col_letter][1:]:
                    cell.number_format = '#,##0'
        # writer.save() ist hier nicht zwingend nötig, da der Kontextmanager das Speichern übernimmt.
    output.seek(0)
    # Sende die Bytes als Download zurück
    return dcc.send_bytes(output.getvalue(), "report_video.xlsx")

# ---------------- bASECHECK ----------------

@app.callback(
    [Output("basecheck-status", "children"),
     Output("basecheck-table", "data"),
     Output("basecheck-table", "columns")],
    Input("calculate-basecheck", "n_clicks"),
    State("mm-dimensions-basecheck", "value")
)
def calculate_basecheck(n_clicks, mm_dims):
    if not n_clicks:
        return "", [], []
    if not mm_dims:
        return "Bitte wählen Sie mindestens eine MM-Dimension aus.", [], []
    
    db_path = "data.db"
    conn = sqlite3.connect(db_path)
    # Lese die Tabelle video (die beide Datensätze enthält)
    df = pd.read_sql("SELECT * FROM data", conn)
    conn.close()
    
    if df.empty:
        return "Die Tabelle video_final ist leer.", [], []
    
    # Filtere nur Zeilen mit hr_basis "Basis" oder "HR"
    df = df[df['hr_basis'].isin(['Basis', 'HR'])]
    
    # Gruppiere nach den ausgewählten Dimensionen plus hr_basis und berechne count(distinct bid)
    grouped = df.groupby(mm_dims + ['hr_basis'], as_index=False)['bid'].nunique()
    
    # Pivot: Index = die in mm_dims gewählten Dimensionen, Spalten = hr_basis
    pivot_df = grouped.pivot_table(index=mm_dims, columns='hr_basis', values='bid', fill_value=0).reset_index()
    
    # Umbenennen der Pivot-Spalten
    pivot_df.rename(columns={'Basis': 'bid_count_basis', 'HR': 'bid_count_hr'}, inplace=True)
    
    # Formatierung: Bid Count als Ganzzahl mit Tausendertrennzeichen
    pivot_df["bid_count_basis"] = pivot_df["bid_count_basis"].apply(lambda x: format(int(x), ",d"))
    pivot_df["bid_count_hr"] = pivot_df["bid_count_hr"].apply(lambda x: format(int(x), ",d"))
    
    # Erstelle Spaltenliste und Daten für die DataTable
    columns = [{"name": col, "id": col} for col in pivot_df.columns]
    data = pivot_df.to_dict("records")
    
    return f"Basecheck: {len(pivot_df)} Gruppen gefunden.", data, columns

#..............................................................NICHT BEWEGTBILD...............................................................



@app.callback(
    [Output("nonvideo-percentages-status", "children"),
     Output("nonvideo-percentages-table", "data"),
     Output("nonvideo-percentages-table", "columns")],
    Input("calculate-percentages2_nbv", "n_clicks"),
    State("mm-dimensions2", "value"),
    State("ea-dimensions2", "value")
)
def calculate_nonvideo_percentages(n_clicks, mm_dims, ea_dims):
    if not n_clicks:
        return "", [], []
    
    # Für die Gesamt-Gruppierung (für ea_hits) verwenden wir alle ausgewählten Felder:
    group_by_all = []
    if mm_dims:
        group_by_all.extend(mm_dims)
    if ea_dims:
        group_by_all.extend(ea_dims)
    
    # Falls keine Dimension gewählt wurde, Fehler zurückgeben.
    if not group_by_all:
        return "Bitte wählen Sie mindestens eine Dimension aus.", [], []
    
    db_path = "data.db"
    conn = sqlite3.connect(db_path)
    # Lese alle Nicht-Video-Daten (Tabelle non_video)
    df_nonvideo = pd.read_sql("SELECT * FROM non_video", conn)
    # Berechne den konstanten Wert: distinct bid über alle Zeilen mit hr_basis = 'Basis'
    distinct_bid_query = "SELECT COUNT(DISTINCT bid) AS bid_count FROM non_video WHERE hr_basis = 'Basis'"
    df_const = pd.read_sql(distinct_bid_query, conn)
    conn.close()
    
    if df_nonvideo.empty:
        return "Die Tabelle non_video ist leer.", [], []
    
    # Filtere nur Zeilen mit hr_basis = "Basis"
    df_basis = df_nonvideo[df_nonvideo['hr_basis'] == 'Basis']
    overall_bid_count = df_basis['bid'].nunique()

    # Hier vorher wurde bisher der Durchschnitt von ave_weighting_factor genommen.
    # Jetzt berechnen wir stattdessen den Durchschnitt des Verhältnisses aus ave_weighted und ave_100:
    df_basis["weight_ratio"] = df_basis.apply(lambda r: r["ave_weighted"] / r["ave_100"] if r["ave_100"] > 0 else 0, axis=1)
    overall_avg_weighting = df_basis["weight_ratio"].mean() * 100

    if ea_dims:
        avg_weighting_df = df_basis.groupby(ea_dims, as_index=False)["weight_ratio"].mean()
        avg_weighting_df.rename(columns={"weight_ratio": "avg_weighting_factor"}, inplace=True)
        avg_weighting_df["avg_weighting_factor"] = avg_weighting_df["avg_weighting_factor"] * 100
    else:
        avg_weighting_df = None


    # Gruppierung 1: bid_mm_kombo – gruppiere nach mm-dimensions2 (Basis)
    if mm_dims:
        bid_mm_df = df_basis.groupby(mm_dims, as_index=False)['bid'].nunique()
        bid_mm_df.rename(columns={'bid': 'bid_mm_kombo'}, inplace=True)
    else:
        bid_mm_df = pd.DataFrame()
    
    # Gruppierung 2: ea_hits – gruppiere nach allen ausgewählten Feldern (Basis)
    ea_hits_df = df_basis.groupby(group_by_all, as_index=False)['bid'].nunique()
    ea_hits_df.rename(columns={'bid': 'ea_hits'}, inplace=True)
    
    # Kennzahl 2: bid_mm_kombo_hr – für HR-Zeilen, gruppiert nach mm-dimensions2
    if mm_dims:
        df_hr = df_nonvideo[df_nonvideo['hr_basis'] == 'HR']
        bid_mm_hr_df = df_hr.groupby(mm_dims, as_index=False)['bid'].nunique()
        bid_mm_hr_df.rename(columns={'bid': 'bid_mm_kombo_hr'}, inplace=True)
    else:
        bid_mm_hr_df = pd.DataFrame()
    
    # Basis-Ergebnis: Alle eindeutigen Kombinationen der in group_by_all gewählten Felder (Basis)
    df_groups = df_basis[group_by_all].drop_duplicates().reset_index(drop=True)
    
    # Merge: Zuerst mit bid_mm_df (auf mm_dims, falls vorhanden)
    if mm_dims:
        df_result = pd.merge(df_groups, bid_mm_df, on=mm_dims, how='left')
    else:
        df_result = df_groups.copy()
    # Dann mit ea_hits_df auf allen group_by_all Feldern
    df_result = pd.merge(df_result, ea_hits_df, on=group_by_all, how='left')
    # Und zusätzlich merge bid_mm_kombo_hr, falls mm_dims gewählt wurden
    if mm_dims:
        df_result = pd.merge(df_result, bid_mm_hr_df, on=mm_dims, how='left')
    
    # Füge den konstanten Gesamtwert hinzu
    df_result["overall_bid_count"] = overall_bid_count

    # Merge avg_weighting_factor basierend auf EA-Dimensionen, falls vorhanden, sonst Gesamtwert
    if ea_dims:
        df_result = pd.merge(df_result, avg_weighting_df, on=ea_dims, how='left')
    else:
        df_result["avg_weighting_factor"] = overall_avg_weighting
    
    # Berechne die Hit Percentage: (ea_hits / bid_mm_kombo) * 100 (nur wenn bid_mm_kombo > 0)
    df_result["hit_percentage"] = df_result.apply(
        lambda row: (row["ea_hits"] / row["bid_mm_kombo"] * 100) if (row.get("bid_mm_kombo", 0) > 0) else 0, axis=1
    )
    
    # Neue Kennzahl: ids_for_HR = bid_mm_kombo_hr * hit_percentage, gerundet auf Ganzzahl
    df_result["ids_for_HR"] = df_result["bid_mm_kombo_hr"] * df_result["hit_percentage"]/100
    df_result["ids_for_HR"] = df_result["ids_for_HR"].apply(lambda x: round(x) if pd.notna(x) else 0)
    
    # Neue Kennzahl: avg_mentions = (Summe der mentions pro Gruppe) / ea_hits, gerundet, mindestens 1.
    # Hier gruppieren wir nach allen ausgewählten Feldern (group_by_all), damit fließen auch die MM-Dimensionen mit ein.
    mentions_df = df_basis.groupby(group_by_all, as_index=False)['mentions'].sum()
    mentions_df.rename(columns={'mentions': 'sum_mentions'}, inplace=True)
    df_result = pd.merge(df_result, mentions_df, on=group_by_all, how='left')
    df_result["avg_mentions"] = df_result.apply(
        lambda row: max(round(row["sum_mentions"] / row["ea_hits"]), 1) if row["ea_hits"] > 0 else 1, axis=1
    )
    
    # Filtere: Nur Gruppen, bei denen ea_hits > 0
    df_result = df_result[df_result["ea_hits"] > 0]
    
    # Formatierung: Wandle numerische Kennzahlen in formatierten Text um.
    if "bid_mm_kombo" in df_result.columns:
        df_result["bid_mm_kombo"] = df_result["bid_mm_kombo"].fillna(0).apply(lambda x: format(int(x), ",d"))
    if "bid_mm_kombo_hr" in df_result.columns:
        df_result["bid_mm_kombo_hr"] = df_result["bid_mm_kombo_hr"].fillna(0).apply(lambda x: format(int(x), ",d"))
    if "ea_hits" in df_result.columns:
        df_result["ea_hits"] = df_result["ea_hits"].fillna(0).apply(lambda x: format(int(x), ",d"))
    df_result["hit_percentage"] = df_result["hit_percentage"].apply(lambda x: f"{x:.2f}")
    df_result["overall_bid_count"] = df_result["overall_bid_count"].fillna(0).apply(lambda x: format(int(x), ",d"))
    df_result["avg_weighting_factor"] = df_result["avg_weighting_factor"].apply(lambda x: f"{x:.2f}")
    df_result["ids_for_HR"] = df_result["ids_for_HR"].apply(lambda x: format(int(x), ",d"))
    df_result["avg_mentions"] = df_result["avg_mentions"].apply(lambda x: format(int(x), ",d"))
    
    columns = [{"name": col, "id": col} for col in df_result.columns]
    data = df_result.to_dict("records")
    
    # Schreibe das Ergebnis in die Datenbank-Tabelle "percent_non_video"
    conn = sqlite3.connect(db_path)
    df_result.to_sql("percent_non_video", conn, if_exists="replace", index=False)
    conn.close()
    
    status_msg = f"Basecheck Non-Video: {len(df_result)} Gruppen gefunden. (Distinct bid Gesamt: {overall_bid_count:,})"
    return status_msg, data, columns

def select_candidate_rows(df_candidates, n_needed, country_field="country", pr_value_field="pr_value", alpha=0.7):
    """
    Wählt n_needed Zeilen aus df_candidates aus, wobei:
      - Die Auswahl proportional zur Anzahl der Zeilen pro country erfolgt.
      - Innerhalb jeder country-Gruppe erfolgt eine gewichtete Auswahl basierend auf pr_value.
        Dabei wird das Gewicht als: weight = alpha * (pr_value_normalized) + (1 - alpha) berechnet.
        So werden hohe pr_value berücksichtigt, aber nicht ausschließlich.
    
    Parameter:
      - df_candidates: DataFrame mit Kandidatenzeilen.
      - n_needed: Gesamtzahl der benötigten Zeilen.
      - country_field: Feldname für die Länderinformation (Standard: "country").
      - pr_value_field: Feldname für den KPI pr_value (Standard: "pr_value").
      - alpha: Gewichtungsfaktor (zwischen 0 und 1) für pr_value.
    
    Rückgabe:
      - Liste von Dictionaries (jede Zeile repräsentiert einen Kandidaten).
    """
    import uuid
    import random
    
    total_candidates = len(df_candidates)
    if total_candidates == 0:
        return []
    
    # Gruppiere nach dem Länderkriterium
    groups = df_candidates.groupby(country_field)
    selected = []
    
    # Proportionale Zuteilung: Aus jeder Gruppe n_country Zeilen auswählen
    for country, group in groups:
        group_size = len(group)
        # Berechne die Anzahl der Zeilen, die aus dieser Gruppe ausgewählt werden sollen
        n_country = round(n_needed * group_size / total_candidates)
        # Wenn die Gruppe nicht leer ist, mindestens 1 auswählen
        if n_country == 0 and group_size > 0:
            n_country = 1
        
        # Berechne Gewichte: Normiere pr_value auf den Maximalwert innerhalb der Gruppe
        pr_values = group[pr_value_field].fillna(0)
        max_pr = pr_values.max()
        if max_pr > 0:
            normalized = pr_values / max_pr
        else:
            normalized = pr_values
        # Gewicht: hohes pr_value soll bevorzugt werden, aber mit einem konstanten Anteil
        weights = alpha * normalized + (1 - alpha)
        
        # Falls genug Zeilen in der Gruppe vorhanden sind, ohne Replacement
        if group_size >= n_country:
            sampled_group = group.sample(n=n_country, weights=weights, replace=False)
        else:
            sampled_group = group.sample(n=n_country, weights=weights, replace=True)
        selected.extend(sampled_group.to_dict(orient="records"))
    
    # Falls insgesamt weniger als n_needed ausgewählt wurden, ergänze zufällig aus dem gesamten DataFrame
    while len(selected) < n_needed:
        extra_row = df_candidates.sample(n=1, replace=False).iloc[0].to_dict()
        extra_row["bid"] = str(uuid.uuid4())
        selected.append(extra_row)
    
    # Falls zu viele Zeilen ausgewählt wurden, zufälliges Trim auf genau n_needed
    if len(selected) > n_needed:
        selected = random.sample(selected, n_needed)
    
    return selected



#--------------non vido HR---------------------------

@app.callback(
    Output("extrapolate-nonvideo-status", "children"),
    Input("extrapolate-nonvideo", "n_clicks"),
    State("mm-dimensions2", "value"),
    State("ea-dimensions2", "value")
)
def extrapolate_nonvideo(n_clicks, mm_dims2, ea_dims2):
    if not n_clicks:
        return ""

    db_path = "data.db"
    try:
        # 1) Lade Aggregations-Tabelle (percent_non_video) und Originaldaten (non_video)
        conn = sqlite3.connect(db_path)
        df_percent = pd.read_sql("SELECT * FROM percent_non_video", conn)
        df_nonvideo = pd.read_sql("SELECT * FROM non_video", conn)
        conn.close()
    except Exception as e:
        return f"Fehler beim Laden der Tabellen: {e}"

    # Sicherheitsabfragen
    if df_percent.empty:
        return "Die Tabelle percent_non_video ist leer."
    if df_nonvideo.empty:
        return "Die Tabelle non_video ist leer."

    # 2) Bestimme MM-Dimensionen
    # Falls keine Auswahl, kannst du einen Fallback definieren oder eine Fehlermeldung zurückgeben
    if not mm_dims2:
        # Beispiel-Fallback oder Abbruch
        return "Keine MM-Dimensionen ausgewählt."

    # 3) Ergebnisliste für neue HR-Zeilen
    result_rows = []

    # 4) Iteriere über jede Zeile in percent_non_video
    for idx, percent_row in df_percent.iterrows():
        # Auslesen der Anzahl an HR-Zeilen, die erstellt werden sollen
        ids_for_hr = 0
        try:
            # Achtung auf Kommas oder Punkt in Strings:
            ids_for_hr = int(float(str(percent_row.get("ids_for_HR", 0)).replace(",", ".")))
        except:
            pass

        # Wenn keine Zeilen gefordert sind, nächste Zeile
        if ids_for_hr <= 0:
            continue

        # Start: Alle non_video-Zeilen mit hr_basis "HR"
        df_candidates = df_nonvideo[df_nonvideo["hr_basis"].str.upper() == "HR"].copy()

        # Für jede in mm-dimensions2 ausgewählte Dimension:
        for dim in mm_dims2:
            # Den in percent_non_video für die Gruppe definierten Wert holen
            val_in_percent = str(percent_row.get(dim, ""))
            # Und nur die Zeilen behalten, bei denen non_video[dim] genau diesem Wert entspricht
            df_candidates = df_candidates[df_candidates[dim].astype(str) == val_in_percent]

        # Falls keine Kandidaten vorhanden, überspringen
        if df_candidates.empty:
            continue

        # Nutze die ausgelagerte Funktion zur Auswahl der Kandidaten
        selected_rows = select_candidate_rows(df_candidates, ids_for_hr, country_field="country", pr_value_field="pr_value", alpha=0.7)




        # 7) Für jede zufällig ausgewählte Zeile: Kopie erstellen und Felder überschreiben
        for cand in selected_rows:
            new_row = cand.copy()

            # (a) Update EA dimensions from percent_row
            if ea_dims2:
                for ea_dim in ea_dims2:
                    if ea_dim in percent_row:
                        new_row[ea_dim] = percent_row[ea_dim]

            # (b) Overwrite mentions (ensure numeric conversion)
            if "avg_mentions" in df_percent.columns:
                try:
                    new_row["mentions"] = int(float(str(percent_row["avg_mentions"]).replace(",", ".")))
                except:
                    new_row["mentions"] = 1  # minimal 1

            # (c) Set ave_100 from pr_value
            new_row['ave_100'] = new_row.get('pr_value')

            # (d) Replace avg_weighting_factor with ave_weighting_factor from percent_row
            if 'avg_weighting_factor' in percent_row:
                try:
                    aw_val = float(str(percent_row['avg_weighting_factor']).replace(",", "."))
                except Exception:
                    aw_val = 0.0
                new_row['ave_weighting_factor'] = aw_val
            else:
                new_row['ave_weighting_factor'] = 0.0

            # (e) Calculate ave_weighted using the new key (ensure pr_value is numeric)
            try:
                pr_val = float(new_row["pr_value"]) if new_row["pr_value"] is not None else 0.0
                new_row["ave_weighted"] = pr_val * (new_row["ave_weighting_factor"] / 100)
            except Exception:
                new_row["ave_weighted"] = 0.0

            # (f) Ensure hr_basis is set to "HR"
            new_row["hr_basis"] = "HR"

            # Append the modified new_row only once
            result_rows.append(new_row)


    # 8) Falls keine neuen Zeilen erstellt wurden -> Meldung
    if not result_rows:
        return "Keine HR-Zeilen extrapoliert."

    # 9) Ergebnis-DatenFrame erstellen und in DB speichern
    df_hr_nonvideo = pd.DataFrame(result_rows)

    try:
        conn = sqlite3.connect(db_path)
        df_hr_nonvideo.to_sql("hr_non_bewegt", conn, if_exists="replace", index=False)
        conn.close()
        # Summiere nochmal die Zeilen für die Meldung
        msg = (
            f"HR Non-Video-Daten extrapoliert: {len(df_hr_nonvideo)} neue Zeilen "
            f"(Summe ids_for_HR = {df_percent['ids_for_HR'].sum()})."
        )
    except Exception as e:
        msg = f"Fehler beim Speichern in 'hr_non_bewegt': {e}"

    return msg


#----------export non-video---------------------

@app.callback(
    Output("download1", "data"),
    Input("export-nonvideo-button", "n_clicks"),
    prevent_initial_call=True
)
def export_nonvideo_to_excel(n_clicks):
    if not n_clicks:
        return None
    db_path = "data.db"
    conn = sqlite3.connect(db_path)
    # Lese beide Tabellen aus der Datenbank:
    df_non_video = pd.read_sql("SELECT * FROM non_video", conn)
    df_hr_non_video = pd.read_sql("SELECT * FROM hr_non_bewegt", conn)
    conn.close()
    
    # Konkatenieren: Zeilen aus non_video und hr_non_video untereinander
    df_export = pd.concat([df_non_video, df_hr_non_video], ignore_index=True)
    
    output = BytesIO()
    # ExcelWriter verwenden – hier wird das Arbeitsblatt explizit "data" genannt.
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name="data")
        workbook = writer.book
        worksheet = writer.sheets["data"]
        
        # Beispiel: Formatierung pro Spalte (anpassen, falls erforderlich)
        for i, col in enumerate(df_export.columns):
            col_letter = get_column_letter(i+1)
            # Beispiel: Zahlenformat für bestimmte Spalten
            if col.lower() in ["pr_value", "ave_100", "ave_weighted"]:
                for cell in worksheet[col_letter][1:]:
                    cell.number_format = '#,##0'
            elif col.lower() in ["broadcasting_time", "visibility", "apt", "start_time_program", "end_time_program", "start_time_item"]:
                for cell in worksheet[col_letter][1:]:
                    cell.number_format = 'h:mm:ss'
    output.seek(0)
    return dcc.send_bytes(output.getvalue(), "nonvideo_report.xlsx")

#------Ergebnisse Non-Video----------------

import plotly.express as px
from dash.dash_table.Format import Format, Group

import plotly.express as px
from dash.dash_table.Format import Format, Group

@app.callback(
    [Output("nonvideo-results-status", "children"),
     Output("nonvideo-results-table", "data"),
     Output("nonvideo-results-table", "columns"),
     Output("nonvideo-pie", "figure")],
    [Input("calculate-results-nonvideo", "n_clicks")],
    [State("mm-dimensions-results", "value"),
     State("ea-dimensions-results", "value"),
     State("hr-basis-filter", "value")]
)
def calculate_nonvideo_results(n_clicks, mm_dims_res, ea_dims_res, hr_basis_filter):
    if not n_clicks:
        return "", [], [], {}
    
    # Laden der Daten
    db_path = "data.db"
    conn = sqlite3.connect(db_path)
    df_non_video = pd.read_sql("SELECT * FROM non_video", conn)
    df_hr_non_video = pd.read_sql("SELECT * FROM hr_non_bewegt", conn)
    conn.close()
    df_combined = pd.concat([df_non_video, df_hr_non_video], ignore_index=True)
    
    # Filter nach hr_basis, falls nicht "all" gewählt
    if hr_basis_filter != "all":
        df_combined = df_combined[df_combined["hr_basis"] == hr_basis_filter]
    
    # Bestimme die Gruppierungsfelder basierend auf der unabhängigen Dimensionsauswahl
    group_by_cols = []
    if mm_dims_res:
        group_by_cols.extend(mm_dims_res)
    if ea_dims_res:
        group_by_cols.extend(ea_dims_res)
    if not group_by_cols:
        return "Bitte wählen Sie mindestens eine Dimension für die Ergebnisberechnung aus.", [], {}, {}
    
    # Aggregation: Summen der Kennzahlen
    agg_df = df_combined.groupby(group_by_cols, as_index=False).agg({
        "mentions": "sum",
        "ave_100": "sum",
        "ave_weighted": "sum"
    })
    
    # Runde die aggregierten Werte und belasse sie als numerische Werte:
    agg_df["Summe mentions"] = agg_df["mentions"].round(0)
    agg_df["Summe ave_100"] = agg_df["ave_100"].round(0)
    agg_df["Summe ave_weighted"] = agg_df["ave_weighted"].round(0)
    agg_df = agg_df.drop(columns=["mentions", "ave_100", "ave_weighted"])
    
    # Definiere die Spalten: Gruppierungsfelder als Text und Kennzahlen als numerisch (mit Tausendertrennzeichen)
    text_columns = [{"name": col, "id": col, "type": "text"} for col in group_by_cols]
    numeric_columns = [
        {"name": "Summe mentions", "id": "Summe mentions", "type": "numeric", "format": Format().group(Group.yes)},
        {"name": "Summe ave_100", "id": "Summe ave_100", "type": "numeric", "format": Format().group(Group.yes)},
        {"name": "Summe ave_weighted", "id": "Summe ave_weighted", "type": "numeric", "format": Format().group(Group.yes)}
    ]
    columns = text_columns + numeric_columns
    data = agg_df.to_dict("records")
    
    # Kreisdiagramm erstellen: Gruppiere die kombinierten Daten nach hr_basis, summiere ave_weighted
    pie_df = df_combined.groupby("hr_basis", as_index=False).agg({"ave_weighted": "sum"})
    fig = px.pie(pie_df, names="hr_basis", values="ave_weighted",
                 title="Verteilung von Summe ave_weighted nach hr_basis")
    
    status_msg = f"Ergebnisse berechnet: {len(agg_df)} Gruppen gefunden."
    return status_msg, data, columns, fig





# ---------------- Main ----------------

if __name__ == '__main__':
    app.run(debug=True)